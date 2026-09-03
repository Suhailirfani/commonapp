import json
from datetime import datetime, timedelta, time
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from apps.tenants.models import Institution, AddOn
from apps.users.models import User
from .models import (
    Competition, Category, Program, Team, Stage, 
    FestDay, Contestant, Participation, GroupParticipation, 
    PointsConfig, ProgramSchedule, CertificateConfig
)

@login_required
def dashboard_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        return redirect('core:judge_dashboard', institution_slug=institution.slug)

    competitions = Competition.objects.filter(institution=institution)
    teams = Team.objects.filter(institution=institution)
    programs = Program.objects.filter(institution=institution)

    if request.user.is_team_leader:
        team = getattr(request.user, 'managed_team', None)
        if not team:
            team = teams.first()

        team_contestants = Contestant.objects.filter(institution=institution, team=team) if team else Contestant.objects.none()

        # Category wise members count for this team
        category_wise_members = []
        base_cats = Category.objects.filter(institution=institution, is_common=False)
        for cat in base_cats:
            cnt = team_contestants.filter(category=cat).count() if team else 0
            category_wise_members.append({'category': cat, 'count': cnt})

        # Announced points & rank from services
        from .services import get_team_standings
        standings = get_team_standings(institution, announced_only=True)
        team_points = 0
        team_rank = "-"
        for s in standings:
            if team and s['team'].id == team.id:
                team_points = s['points']
                team_rank = s['position']
                break

        # Assigned programs count for this team's contestants
        assigned_single_count = Participation.objects.filter(contestant__team=team).values('program').distinct().count() if team else 0
        assigned_group_count = GroupParticipation.objects.filter(team=team).values('program').distinct().count() if team else 0
        total_assignments = assigned_single_count + assigned_group_count

        context = {
            'institution': institution,
            'is_team_leader_dashboard': True,
            'team': team,
            'total_members_count': team_contestants.count(),
            'category_wise_members': category_wise_members,
            'team_points': team_points,
            'team_rank': team_rank,
            'total_assignments': total_assignments,
            'programs_count': programs.count(),
        }
        return render(request, 'core/dashboard.html', context)

    contestants = Contestant.objects.filter(institution=institution)

    context = {
        'institution': institution,
        'competitions_count': competitions.count(),
        'teams_count': teams.count(),
        'programs_count': programs.count(),
        'contestants_count': contestants.count(),
        'recent_competitions': competitions[:5],
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def competition_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    return render(request, 'core/competition_list.html', {'institution': institution, 'competitions': competitions})


@login_required
def competition_create_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        name = request.POST.get('name')
        comp_type = request.POST.get('type')
        year = request.POST.get('year', 2026)
        logo = request.FILES.get('logo')
        name_image = request.FILES.get('name_image')
        max_single = int(request.POST.get('max_single_programs_per_contestant', 0) or 0)
        max_group = int(request.POST.get('max_group_programs_per_contestant', 0) or 0)
        max_total = int(request.POST.get('max_total_programs_per_contestant', 0) or 0)
        max_team_single = int(request.POST.get('max_team_participants_per_single_program', 0) or 0)
        max_team_group = int(request.POST.get('max_team_entries_per_group_program', 0) or 0)
        Competition.objects.create(
            institution=institution,
            name=name,
            type=comp_type,
            year=year,
            logo=logo,
            name_image=name_image,
            max_single_programs_per_contestant=max_single,
            max_group_programs_per_contestant=max_group,
            max_total_programs_per_contestant=max_total,
            max_team_participants_per_single_program=max_team_single,
            max_team_entries_per_group_program=max_team_group,
        )
        messages.success(request, f"Fest '{name}' created successfully!")
        return redirect('core:competition_list', institution_slug=institution.slug)
    return render(request, 'core/competition_create.html', {'institution': institution})


@login_required
def competition_edit_view(request, institution_slug, comp_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    comp = get_object_or_404(Competition, id=comp_id, institution=institution)
    if request.method == 'POST':
        name = request.POST.get('name')
        comp_type = request.POST.get('type')
        year = request.POST.get('year', 2026)
        max_single = int(request.POST.get('max_single_programs_per_contestant', 0) or 0)
        max_group = int(request.POST.get('max_group_programs_per_contestant', 0) or 0)
        max_total = int(request.POST.get('max_total_programs_per_contestant', 0) or 0)
        max_team_single = int(request.POST.get('max_team_participants_per_single_program', 0) or 0)
        max_team_group = int(request.POST.get('max_team_entries_per_group_program', 0) or 0)

        comp.name = name
        comp.type = comp_type
        comp.year = year
        comp.max_single_programs_per_contestant = max_single
        comp.max_group_programs_per_contestant = max_group
        comp.max_total_programs_per_contestant = max_total
        comp.max_team_participants_per_single_program = max_team_single
        comp.max_team_entries_per_group_program = max_team_group

        if request.POST.get('clear_logo') == '1':
            comp.logo = None
        elif 'logo' in request.FILES:
            comp.logo = request.FILES['logo']

        if request.POST.get('clear_name_image') == '1':
            comp.name_image = None
        elif 'name_image' in request.FILES:
            comp.name_image = request.FILES['name_image']

        if request.POST.get('clear_custom_result_template') == '1':
            if comp.custom_result_template:
                comp.custom_result_template.delete(save=False)
            comp.custom_result_template = None
        elif 'custom_result_template' in request.FILES:
            if comp.custom_result_template:
                comp.custom_result_template.delete(save=False)
            comp.custom_result_template = request.FILES['custom_result_template']

        comp.save()
        messages.success(request, f"Fest '{name}' updated successfully!")
        return redirect('core:competition_list', institution_slug=institution.slug)
    return render(request, 'core/competition_edit.html', {'institution': institution, 'comp': comp})


@login_required
def competition_delete_view(request, institution_slug, comp_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    comp = get_object_or_404(Competition, id=comp_id, institution=institution)

    programs_count = comp.programs.count()
    contestants_count = comp.contestants.count()
    categories_count = comp.categories.count()
    teams_count = comp.teams.count()

    if request.method == 'POST':
        name = comp.name
        comp.delete()
        messages.success(request, f"Fest '{name}' and all associated programs, contestants, categories & teams have been permanently deleted.")
        return redirect('core:competition_list', institution_slug=institution.slug)

    context = {
        'institution': institution,
        'comp': comp,
        'programs_count': programs_count,
        'contestants_count': contestants_count,
        'categories_count': categories_count,
        'teams_count': teams_count,
    }
    return render(request, 'core/competition_delete_confirm.html', context)


@login_required
def category_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    
    if request.method == 'POST':
        if not institution.allows_category_management:
            messages.error(request, "🔒 Category creation and editing is currently locked in Settings.")
            return redirect('core:category_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        is_common = request.POST.get('is_common') == '1'
        inc_cat_ids = request.POST.getlist('included_categories[]')

        comp = Competition.objects.filter(id=comp_id, institution=institution).first()
        if not comp:
            comp = competitions.first()

        if comp and name:
            cat = Category.objects.create(
                institution=institution,
                competition=comp,
                name=name,
                description=description,
                is_common=is_common
            )
            if is_common and inc_cat_ids:
                inc_cats = Category.objects.filter(id__in=inc_cat_ids, institution=institution)
                cat.included_categories.set(inc_cats)

            cat_type = "Common Category" if is_common else "Base Category"
            messages.success(request, f"{cat_type} '{name}' created successfully!")
            return redirect('core:category_list', institution_slug=institution.slug)

    categories = Category.objects.filter(institution=institution).prefetch_related('included_categories')
    base_categories = categories.filter(is_common=False)

    return render(request, 'core/category_list.html', {
        'institution': institution,
        'categories': categories,
        'base_categories': base_categories,
        'competitions': competitions
    })


@login_required
def program_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = list(Competition.objects.filter(institution=institution))
    categories = list(Category.objects.filter(institution=institution))

    if request.method == 'POST':
        if not institution.allows_program_management:
            messages.error(request, "🔒 Program creation, editing, and import are currently locked in Settings.")
            return redirect('core:program_list', institution_slug=institution.slug)

        action = request.POST.get('action', '')

        # Action 1: Create Single Program
        if action == 'create_single_program':
            comp_id = request.POST.get('competition_id')
            cat_id = request.POST.get('category_id')
            name = request.POST.get('name', '').strip()
            is_group = request.POST.get('is_group') in ['on', '1', 'true']
            p_type = request.POST.get('program_type', 'STAGE')
            p_mode = request.POST.get('presentation_mode', 'SEQUENTIAL')
            duration = request.POST.get('duration_per_participant', 5)

            comp = Competition.objects.filter(id=comp_id, institution=institution).first()
            if not comp and competitions:
                comp = competitions[0]
            cat = Category.objects.filter(id=cat_id, institution=institution).first()

            if comp and cat and name:
                Program.objects.create(
                    institution=institution,
                    competition=comp,
                    category=cat,
                    name=name,
                    is_group=is_group,
                    program_type=p_type,
                    presentation_mode=p_mode,
                    duration_per_participant=int(duration) if str(duration).isdigit() else 5
                )
                messages.success(request, f"Program '{name}' created successfully!")
                return redirect(f"{reverse('core:program_list', kwargs={'institution_slug': institution.slug})}?tab=list")
            else:
                messages.error(request, "Failed to create program. Program Name, Fest and Category are required.")

        # Action 2: WhatsApp Text Bulk Import
        elif action == 'whatsapp_import':
            wa_text = request.POST.get('whatsapp_text', '').strip()
            default_cat_id = request.POST.get('default_category_id')
            comp_id = request.POST.get('competition_id')

            comp = Competition.objects.filter(id=comp_id, institution=institution).first()
            if not comp and competitions:
                comp = competitions[0]

            default_cat = Category.objects.filter(id=default_cat_id, institution=institution).first()
            current_cat = default_cat

            import re
            created_count = 0
            for line in wa_text.splitlines():
                line_str = line.strip()
                if not line_str:
                    continue

                # Skip titles/headers
                clean_upper = line_str.upper().strip('*#=- ')
                if clean_upper in ['PROGRAM LIST', 'PROGRAMS LIST', 'PROGRAMS', 'ITEMS LIST', 'EVENT LIST', 'SCHEDULE']:
                    continue

                if 'category' in line_str.lower() or line_str.lower().startswith('cat:'):
                    cat_name = re.sub(r'^(category|cat)[\s\:\-]*', '', line_str, flags=re.IGNORECASE).strip('* ').strip()
                    if cat_name:
                        cat_obj = Category.objects.filter(institution=institution, name__iexact=cat_name).first()
                        if not cat_obj and comp:
                            cat_obj = Category.objects.create(institution=institution, competition=comp, name=cat_name)
                        if cat_obj:
                            current_cat = cat_obj
                    continue

                clean_line = re.sub(r'^\d+[\.\)\-]*\s*', '', line_str).strip('* ').strip()
                if not clean_line or len(clean_line) < 2:
                    continue

                parts = [p.strip() for p in clean_line.split('-') if p.strip()]
                prog_name = parts[0]

                is_group = bool(re.search(r'\b(GROUP|TEAM)\b', clean_line, re.IGNORECASE))
                is_offstage = bool(re.search(r'\b(OFF\s*STAGE|OFFSTAGE|WRITTEN|ART)\b', clean_line, re.IGNORECASE))
                p_type = 'OFF_STAGE' if is_offstage else 'STAGE'

                if not current_cat:
                    current_cat = Category.objects.filter(institution=institution).first()
                    if not current_cat and comp:
                        current_cat = Category.objects.create(institution=institution, competition=comp, name='General Category')

                if comp and current_cat and prog_name:
                    Program.objects.create(
                        institution=institution,
                        competition=comp,
                        category=current_cat,
                        name=prog_name,
                        is_group=is_group,
                        program_type=p_type
                    )
                    created_count += 1

            if created_count > 0:
                messages.success(request, f"Successfully imported {created_count} program(s) from WhatsApp text!")
                return redirect(f"{reverse('core:program_list', kwargs={'institution_slug': institution.slug})}?tab=list")
            else:
                messages.error(request, "Could not parse any valid programs from the provided WhatsApp text. Please check format.")

    programs = Program.objects.filter(institution=institution).select_related('category', 'competition').order_by('category__name', 'name')
    return render(request, 'core/program_list.html', {
        'institution': institution,
        'programs': programs,
        'competitions': competitions,
        'categories': categories,
    })


@login_required
def program_create_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution)
    
    if request.method == 'POST':
        if not institution.allows_program_management:
            messages.error(request, "🔒 Program creation, editing, and import are currently locked in Settings.")
            return redirect('core:program_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        cat_id = request.POST.get('category_id')
        name = request.POST.get('name')
        is_group = request.POST.get('is_group') == 'on'
        p_type = request.POST.get('program_type', 'STAGE')
        p_mode = request.POST.get('presentation_mode', 'SEQUENTIAL')
        duration = request.POST.get('duration_per_participant', 5)
        comp = get_object_or_404(Competition, id=comp_id, institution=institution)
        cat = get_object_or_404(Category, id=cat_id, institution=institution)
        max_team = int(request.POST.get('max_participants_per_team', 0) or 0)

        Program.objects.create(
            institution=institution,
            competition=comp,
            category=cat,
            name=name,
            is_group=is_group,
            program_type=p_type,
            presentation_mode=p_mode,
            duration_per_participant=duration,
            max_participants_per_team=max_team
        )
        messages.success(request, f"Program '{name}' created successfully!")
        return redirect('core:program_list', institution_slug=institution.slug)

    return render(request, 'core/program_create.html', {
        'institution': institution, 
        'competitions': competitions, 
        'categories': categories
    })


@login_required
def program_batch_create_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution)

    if request.method == 'POST':
        comp_ids = request.POST.getlist('competition_id[]')
        cat_ids = request.POST.getlist('category_id[]')
        names = request.POST.getlist('name[]')
        p_types = request.POST.getlist('program_type[]')
        is_groups = request.POST.getlist('is_group[]')
        p_modes = request.POST.getlist('presentation_mode[]')
        durations = request.POST.getlist('duration[]')
        team_limits = request.POST.getlist('max_participants_per_team[]')

        created_count = 0
        for i in range(len(names)):
            prog_name = names[i].strip() if i < len(names) else ''
            if not prog_name:
                continue

            comp_id = comp_ids[i] if i < len(comp_ids) else None
            cat_id = cat_ids[i] if i < len(cat_ids) else None
            p_type = p_types[i] if i < len(p_types) else 'STAGE'
            is_group_val = (is_groups[i] == '1' or is_groups[i] == 'true' or is_groups[i] == 'on') if i < len(is_groups) else False
            p_mode = p_modes[i] if i < len(p_modes) else 'SEQUENTIAL'
            duration_val = int(durations[i]) if i < len(durations) and str(durations[i]).isdigit() else 5
            max_team_val = int(team_limits[i]) if i < len(team_limits) and str(team_limits[i]).isdigit() else 0

            comp = Competition.objects.filter(id=comp_id, institution=institution).first()
            cat = Category.objects.filter(id=cat_id, institution=institution).first()

            if comp and cat:
                Program.objects.create(
                    institution=institution,
                    competition=comp,
                    category=cat,
                    name=prog_name,
                    is_group=is_group_val,
                    program_type=p_type,
                    presentation_mode=p_mode,
                    duration_per_participant=duration_val,
                    max_participants_per_team=max_team_val
                )
                created_count += 1

        messages.success(request, f"Successfully created {created_count} programs in batch!")
        return redirect('core:program_list', institution_slug=institution.slug)

    return render(request, 'core/program_batch_create.html', {
        'institution': institution,
        'competitions': competitions,
        'categories': categories,
    })


@login_required
def program_download_template_view(request, institution_slug):
    import io, openpyxl
    from django.http import HttpResponse

    institution = get_object_or_404(Institution, slug=institution_slug)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programs Template"

    headers = ["Competition Name", "Category Name", "Program Name", "Program Type (STAGE/OFF_STAGE)", "Format (SINGLE/GROUP)", "Presentation Mode (SEQ/SIM)", "Duration Mins"]
    ws.append(headers)

    sample_comp = Competition.objects.filter(institution=institution).first()
    comp_name = sample_comp.name if sample_comp else "Mueeniyya Grand Fest 2026"

    sample_rows = [
        [comp_name, "Junior Category", "Quran Recitation", "STAGE", "SINGLE", "SEQ", 5],
        [comp_name, "Junior Category", "Group Song", "STAGE", "GROUP", "SEQ", 10],
        [comp_name, "Senior Category", "Pencil Drawing", "OFF_STAGE", "SINGLE", "SIM", 30],
        [comp_name, "Senior Category", "Elocution English", "STAGE", "SINGLE", "SEQ", 7],
    ]
    for row in sample_rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{institution.slug}_programs_template.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def program_bulk_upload_view(request, institution_slug):
    import openpyxl
    institution = get_object_or_404(Institution, slug=institution_slug)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        if not institution.allows_program_management:
            messages.error(request, "🔒 Program creation, editing, and import are currently locked in Settings.")
            return redirect('core:program_list', institution_slug=institution.slug)

        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) <= 1:
                messages.error(request, "Uploaded Excel file contains no data rows.")
                return redirect('core:program_list', institution_slug=institution.slug)

            imported_count = 0
            for row in rows[1:]:
                if not row or not any(row):
                    continue

                comp_name = str(row[0]).strip() if len(row) > 0 and row[0] else "Main Fest"
                cat_name = str(row[1]).strip() if len(row) > 1 and row[1] else "General"
                prog_name = str(row[2]).strip() if len(row) > 2 and row[2] else None

                if not prog_name:
                    continue

                p_type_str = str(row[3]).strip().upper() if len(row) > 3 and row[3] else "STAGE"
                format_str = str(row[4]).strip().upper() if len(row) > 4 and row[4] else "SINGLE"
                
                # Check column 5 for Presentation Mode vs Duration
                p_mode_str = str(row[5]).strip().upper() if len(row) > 5 and row[5] else "SEQ"
                if p_mode_str.isdigit():
                    duration_val = int(p_mode_str)
                    presentation_mode = 'SEQUENTIAL'
                else:
                    duration_val = int(row[6]) if len(row) > 6 and str(row[6]).isdigit() else 5
                    presentation_mode = 'SIMULTANEOUS' if any(k in p_mode_str for k in ['SIMULTANEOUS', 'SIM', 'ALL', 'WRITTEN']) else 'SEQUENTIAL'

                comp, _ = Competition.objects.get_or_create(
                    institution=institution,
                    name=comp_name,
                    defaults={'type': 'ON', 'year': 2026}
                )

                cat, _ = Category.objects.get_or_create(
                    institution=institution,
                    competition=comp,
                    name=cat_name
                )

                program_type = 'OFF_STAGE' if 'OFF' in p_type_str else 'STAGE'
                is_group = True if (format_str in ['GROUP', 'G', 'YES', 'TRUE', '1'] or 'GROUP' in format_str or format_str == 'G') else False

                Program.objects.create(
                    institution=institution,
                    competition=comp,
                    category=cat,
                    name=prog_name,
                    program_type=program_type,
                    is_group=is_group,
                    presentation_mode=presentation_mode,
                    duration_per_participant=duration_val
                )
                imported_count += 1

            messages.success(request, f"Excel import successful! Imported {imported_count} programs.")
        except Exception as e:
            messages.error(request, f"Failed to parse Excel file: {str(e)}")

    return redirect('core:program_list', institution_slug=institution.slug)


@login_required
def team_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    team_leaders = User.objects.filter(institution=institution, role='TEAM_LEADER')
    
    if request.method == 'POST':
        if not institution.allows_team_management:
            messages.error(request, "🔒 Team creation and editing is currently locked in Settings.")
            return redirect('core:team_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        name = request.POST.get('name')
        code_letter = request.POST.get('code_letter', '').strip().upper()
        leader_id = request.POST.get('leader_id')
        logo = request.FILES.get('logo')

        comp = Competition.objects.filter(id=comp_id, institution=institution).first()
        if not comp:
            comp = competitions.first()

        if comp and name:
            team = Team.objects.create(
                institution=institution,
                competition=comp,
                name=name,
                code_letter=code_letter,
                logo=logo
            )
            if leader_id:
                leader = User.objects.filter(id=leader_id, institution=institution, role='TEAM_LEADER').first()
                if leader:
                    Team.objects.filter(institution=institution, user=leader).exclude(id=team.id).update(user=None)
                    team.user = leader
                    team.save()

            messages.success(request, f"Team '{name}' created successfully!")
            return redirect('core:team_list', institution_slug=institution.slug)
        else:
            messages.error(request, "Failed to create team. Name and competition are required.")

    teams = Team.objects.filter(institution=institution).select_related('competition', 'user')

    return render(request, 'core/team_list.html', {
        'institution': institution,
        'teams': teams,
        'competitions': competitions,
        'team_leaders': team_leaders
    })


@login_required
def contestant_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    contestants = Contestant.objects.filter(institution=institution).select_related('team', 'category')
    managed_team = None

    if request.user.is_team_leader:
        managed_team = getattr(request.user, 'managed_team', None)
        if managed_team:
            contestants = contestants.filter(team=managed_team)
        else:
            contestants = Contestant.objects.none()

    has_contestant_addon = institution.has_add_on('contestant-user-creation')

    return render(request, 'core/contestant_list.html', {
        'institution': institution, 
        'contestants': contestants,
        'managed_team': managed_team,
        'has_contestant_addon': has_contestant_addon,
    })


@login_required
def contestant_create_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution, is_common=False)
    teams = Team.objects.filter(institution=institution)

    if request.method == 'POST':
        if not institution.allows_contestant_registration:
            messages.error(request, "🔒 Contestant registration, editing, and bulk upload are currently locked in Settings.")
            return redirect('core:contestant_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        team_id = request.POST.get('team_id')
        cat_id = request.POST.get('category_id')
        name = request.POST.get('name')
        wa_num = request.POST.get('whatsapp_number', '').strip()

        comp = get_object_or_404(Competition, id=comp_id, institution=institution)
        team = get_object_or_404(Team, id=team_id, institution=institution)
        cat = get_object_or_404(Category, id=cat_id, institution=institution)

        if cat.is_common:
            messages.error(request, f"Direct contestant registration to Combined Category '{cat.name}' is not allowed. Please select a Base Category.")
            return redirect('core:contestant_create', institution_slug=institution.slug)

        c = Contestant.objects.create(
            institution=institution,
            competition=comp,
            team=team,
            category=cat,
            name=name,
            whatsapp_number=wa_num
        )
        messages.success(request, f"Contestant #{c.chest_no} '{c.name}' registered successfully!")
        return redirect('core:contestant_list', institution_slug=institution.slug)

    return render(request, 'core/contestant_create.html', {
        'institution': institution,
        'competitions': competitions,
        'categories': categories,
        'teams': teams
    })


@login_required
def contestant_batch_create_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution, is_common=False)
    teams = Team.objects.filter(institution=institution)

    if request.method == 'POST':
        if not institution.allows_contestant_registration:
            messages.error(request, "🔒 Contestant registration, editing, and bulk upload are currently locked in Settings.")
            return redirect('core:contestant_list', institution_slug=institution.slug)

        comp_ids = request.POST.getlist('competition_id[]')
        team_ids = request.POST.getlist('team_id[]')
        cat_ids = request.POST.getlist('category_id[]')
        names = request.POST.getlist('name[]')
        wa_numbers = request.POST.getlist('whatsapp_number[]')

        registered_count = 0
        for i in range(len(names)):
            c_name = names[i].strip() if i < len(names) else ''
            if not c_name:
                continue

            comp_id = comp_ids[i] if i < len(comp_ids) else None
            team_id = team_ids[i] if i < len(team_ids) else None
            cat_id = cat_ids[i] if i < len(cat_ids) else None
            wa_num = wa_numbers[i].strip() if i < len(wa_numbers) else ''

            comp = Competition.objects.filter(id=comp_id, institution=institution).first()
            team = Team.objects.filter(id=team_id, institution=institution).first()
            cat = Category.objects.filter(id=cat_id, institution=institution, is_common=False).first()

            if comp and team and cat:
                Contestant.objects.create(
                    institution=institution,
                    competition=comp,
                    team=team,
                    category=cat,
                    name=c_name,
                    whatsapp_number=wa_num
                )
                registered_count += 1

        messages.success(request, f"Successfully registered {registered_count} contestants in batch!")
        return redirect('core:contestant_list', institution_slug=institution.slug)

    return render(request, 'core/contestant_batch_create.html', {
        'institution': institution,
        'competitions': competitions,
        'categories': categories,
        'teams': teams,
    })


@login_required
def contestant_download_template_view(request, institution_slug):
    import io, openpyxl
    from django.http import HttpResponse

    institution = get_object_or_404(Institution, slug=institution_slug)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contestants Template"

    headers = ["Fest Name", "Team Name", "Category Name", "Contestant Name", "Chest No (Optional)", "WhatsApp Number (Optional)"]
    ws.append(headers)

    sample_comp = Competition.objects.filter(institution=institution).first()
    comp_name = sample_comp.name if sample_comp else "Mueeniyya Grand Fest 2026"

    sample_rows = [
        [comp_name, "Red House Alpha", "Junior Category", "Ahmad Bilal", 1001, "9876543210"],
        [comp_name, "Blue House Titans", "Junior Category", "Zayd Haris", 1002, "9876543211"],
        [comp_name, "Green Gladiators", "Senior Category", "Hamza Ali", 1003, ""],
        [comp_name, "Red House Alpha", "Senior Category", "Umar Farooq", "", "9876543212"],
    ]
    for row in sample_rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{institution.slug}_contestants_template.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def contestant_bulk_upload_view(request, institution_slug):
    import openpyxl
    institution = get_object_or_404(Institution, slug=institution_slug)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        if not institution.allows_contestant_registration:
            messages.error(request, "🔒 Contestant registration, editing, and bulk upload are currently locked in Settings.")
            return redirect('core:contestant_list', institution_slug=institution.slug)

        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) <= 1:
                messages.error(request, "Uploaded Excel file contains no data rows.")
                return redirect('core:contestant_list', institution_slug=institution.slug)

            imported_count = 0
            for row in rows[1:]:
                if not row or not any(row):
                    continue

                comp_name = str(row[0]).strip() if len(row) > 0 and row[0] else "Main Fest"
                team_name = str(row[1]).strip() if len(row) > 1 and row[1] else "General Team"
                cat_name = str(row[2]).strip() if len(row) > 2 and row[2] else "General"
                c_name = str(row[3]).strip() if len(row) > 3 and row[3] else None

                if not c_name:
                    continue

                chest_no_raw = row[4] if len(row) > 4 and row[4] else None
                chest_no = int(chest_no_raw) if chest_no_raw and str(chest_no_raw).isdigit() else None

                wa_num_raw = row[5] if len(row) > 5 and row[5] else ""
                wa_num = str(wa_num_raw).strip() if wa_num_raw else ""

                comp, _ = Competition.objects.get_or_create(
                    institution=institution,
                    name=comp_name,
                    defaults={'type': 'ON', 'year': 2026}
                )

                team, _ = Team.objects.get_or_create(
                    institution=institution,
                    competition=comp,
                    name=team_name
                )

                cat, _ = Category.objects.get_or_create(
                    institution=institution,
                    competition=comp,
                    name=cat_name
                )

                if chest_no:
                    if Contestant.objects.filter(institution=institution, competition=comp, chest_no=chest_no).exists():
                        from apps.core.services import get_next_chest_number
                        chest_no = get_next_chest_number(cat)

                Contestant.objects.create(
                    institution=institution,
                    competition=comp,
                    team=team,
                    category=cat,
                    name=c_name,
                    chest_no=chest_no,
                    whatsapp_number=wa_num
                )
                imported_count += 1

            messages.success(request, f"Excel import successful! Registered {imported_count} contestants.")
        except Exception as e:
            messages.error(request, f"Failed to parse Excel file: {str(e)}")

    return redirect('core:contestant_list', institution_slug=institution.slug)


@login_required
def judge_dashboard_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if not request.user.is_judge and not (request.user.is_developer or request.user.is_institution_admin):
        return redirect('core:dashboard', institution_slug=institution.slug)

    programs = Program.objects.filter(institution=institution).select_related('category', 'competition')

    if request.user.is_judge:
        judge_comps = request.user.assigned_competitions.all()
        judge_progs = request.user.assigned_programs.all()
        if judge_comps.exists() and judge_progs.exists():
            programs = programs.filter(Q(competition__in=judge_comps) | Q(id__in=judge_progs))
        elif judge_comps.exists():
            programs = programs.filter(competition__in=judge_comps)
        elif judge_progs.exists():
            programs = programs.filter(id__in=judge_progs)
        else:
            programs = Program.objects.none()

    program_list = []
    completed_count = 0
    pending_count = 0

    for p in programs:
        if p.is_group:
            marked = GroupParticipation.objects.filter(program=p, marks__isnull=False).exists()
            total_parts = GroupParticipation.objects.filter(program=p).count()
        else:
            marked = Participation.objects.filter(program=p, marks__isnull=False).exists()
            total_parts = Participation.objects.filter(program=p).count()

        if marked:
            completed_count += 1
        else:
            pending_count += 1

        program_list.append({
            'program': p,
            'is_completed': marked,
            'total_participants': total_parts
        })

    return render(request, 'core/judge_dashboard.html', {
        'institution': institution,
        'program_list': program_list,
        'total_assigned': len(program_list),
        'completed_count': completed_count,
        'pending_count': pending_count,
    })


@login_required
def scoring_program_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    programs = Program.objects.filter(institution=institution).select_related('category', 'competition')
    categories = Category.objects.filter(institution=institution)

    if request.user.is_judge:
        judge_comps = request.user.assigned_competitions.all()
        judge_progs = request.user.assigned_programs.all()
        if judge_comps.exists() and judge_progs.exists():
            programs = programs.filter(Q(competition__in=judge_comps) | Q(id__in=judge_progs))
        elif judge_comps.exists():
            programs = programs.filter(competition__in=judge_comps)
        elif judge_progs.exists():
            programs = programs.filter(id__in=judge_progs)
        else:
            programs = Program.objects.none()

    for p in programs:
        if p.is_group:
            p.has_marks = GroupParticipation.objects.filter(program=p, marks__isnull=False).exists()
        else:
            p.has_marks = Participation.objects.filter(program=p, marks__isnull=False).exists()

    return render(request, 'core/scoring_program_list.html', {
        'institution': institution,
        'programs': programs,
        'categories': categories,
    })


@login_required
def mark_entry_matrix_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if request.user.is_judge:
        judge_comps = request.user.assigned_competitions.all()
        judge_progs = request.user.assigned_programs.all()
        is_assigned_comp = judge_comps.filter(id=program.competition_id).exists()
        is_assigned_prog = judge_progs.filter(id=program.id).exists()
        if not (is_assigned_comp or is_assigned_prog):
            messages.error(request, "Permission Denied: You are not assigned to score this program/competition.")
            return redirect('core:scoring_program_list', institution_slug=institution.slug)

    if program.is_group:
        participations = GroupParticipation.objects.filter(program=program).select_related('team')
        has_marks = GroupParticipation.objects.filter(program=program, marks__isnull=False).exists()
    else:
        participations = Participation.objects.filter(program=program).select_related('contestant', 'contestant__team')
        has_marks = Participation.objects.filter(program=program, marks__isnull=False).exists()

    if request.method == 'POST':
        # Action 1: Update Judge Settings (Count & Max Marks per judge)
        if 'action' in request.POST and request.POST.get('action') == 'update_judge_settings':
            try:
                new_jc = int(request.POST.get('judge_count', 1))
                new_max = int(request.POST.get('max_marks_per_judge', 100))
                if new_jc > 0 and new_max > 0:
                    program.judge_count = new_jc
                    program.max_marks_per_judge = new_max
                    program.save()
                    messages.success(request, f"Updated judge configuration: {new_jc} judge(s), Max {new_max} marks per judge.")
                else:
                    messages.error(request, "Judge count and max marks per judge must be greater than 0.")
            except ValueError:
                messages.error(request, "Invalid judge configuration values.")
            return redirect('core:mark_entry_matrix', institution_slug=institution.slug, program_id=program.id)

        # Action 2: Save Marks Entry Matrix
        part_ids = set()
        for key in request.POST.keys():
            if key.startswith('code_letter_'):
                part_ids.add(key.replace('code_letter_', ''))
            elif key.startswith('marks_'):
                part_ids.add(key.replace('marks_', ''))
            elif key.startswith('j_') and '_marks_' in key:
                parts = key.split('_marks_')
                if len(parts) == 2:
                    part_ids.add(parts[1])

        jc = program.judge_count or 1
        max_per_judge = program.max_marks_per_judge or 100

        for part_id in part_ids:
            code_vals = [v.strip() for v in request.POST.getlist(f'code_letter_{part_id}') if v.strip() != '']
            code = code_vals[0] if code_vals else ''

            j_dict = {}
            raw_sum = 0
            has_judge_entry = False
            active_judge_count = 0

            for j_num in range(1, jc + 1):
                j_vals = [v.strip() for v in request.POST.getlist(f'j_{j_num}_marks_{part_id}') if v.strip() != '']
                if j_vals:
                    try:
                        val = float(j_vals[0])
                        j_dict[str(j_num)] = val
                        raw_sum += val
                        has_judge_entry = True
                        active_judge_count += 1
                    except ValueError:
                        j_dict[str(j_num)] = None

            if not has_judge_entry:
                marks_vals = [v.strip() for v in request.POST.getlist(f'marks_{part_id}') if v.strip() != '']
                if marks_vals:
                    try:
                        val = float(marks_vals[0])
                        j_dict['1'] = val
                        raw_sum = val
                        has_judge_entry = True
                        active_judge_count = 1
                    except ValueError:
                        pass

            converted_marks_100 = None
            if has_judge_entry and active_judge_count > 0:
                total_max_possible = active_judge_count * max_per_judge
                if total_max_possible > 0:
                    converted_score = (raw_sum / total_max_possible) * 100.0
                    converted_marks_100 = int(round(converted_score))

            if program.is_group:
                p = GroupParticipation.objects.filter(id=part_id, program=program).first()
            else:
                p = Participation.objects.filter(id=part_id, program=program).first()

            if p:
                p.code_letter = code
                p.marks = converted_marks_100
                p.judge_marks = j_dict
                p.save()

        from .services import calculate_program_results
        calculate_program_results(program)

        messages.success(request, f"Marks saved and converted to /100 scale for '{program.name}'!")
        return redirect('core:mark_entry_matrix', institution_slug=institution.slug, program_id=program.id)

    judge_range = list(range(1, (program.judge_count or 1) + 1))

    return render(request, 'core/mark_entry_matrix.html', {
        'institution': institution,
        'program': program,
        'participations': participations,
        'judge_range': judge_range,
        'has_marks': has_marks,
    })


@login_required
def manage_results_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    return render(request, 'core/manage_results.html', {
        'institution': institution,
    })


@login_required
def judge_management_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)

    if not (request.user.is_developer or request.user.is_institution_admin or request.user.role in ['SUB_ADMIN', 'INSTITUTION_ADMIN']):
        messages.error(request, "Access Restricted: Only Institution Admins can manage judge assignments.")
        return redirect('core:dashboard', institution_slug=institution.slug)

    from apps.users.models import User
    judges = User.objects.filter(
        Q(institution=institution) | Q(role='JUDGE'),
        role__in=['JUDGE', 'TABULATOR', 'SUB_ADMIN']
    ).distinct().prefetch_related('assigned_programs', 'assigned_programs__category').order_by('first_name', 'username')

    comp = Competition.objects.filter(institution=institution, is_active=True).first()
    if not comp:
        comp = Competition.objects.filter(institution=institution).first()

    categories = list(Category.objects.filter(institution=institution).order_by('id'))
    all_programs = list(Program.objects.filter(institution=institution).select_related('category').order_by('category__name', 'name'))

    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Action A: Quick Create New Judge User Account
        if action == 'create_judge_user':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            name = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()

            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' already exists. Please choose another username.")
                else:
                    User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=name,
                        phone=phone,
                        role='JUDGE',
                        institution=institution,
                        is_approved=True
                    )
                    messages.success(request, f"Successfully created Judge User '{username}' ({name or username})!")
            else:
                messages.error(request, "Username and Password are required to create a Judge account.")
            return redirect('core:judge_management', institution_slug=institution.slug)

        # Action B: Save Assigned Programs for a Specific Judge
        if action == 'update_judge_programs':
            judge_user_id = request.POST.get('judge_user_id')
            j_user = User.objects.filter(id=judge_user_id, institution=institution).first()
            if j_user:
                prog_ids = request.POST.getlist('judge_assigned_programs')
                clean_prog_ids = [int(p) for p in prog_ids if p and p.strip().isdigit()]
                valid_progs = Program.objects.filter(id__in=clean_prog_ids, institution=institution)
                j_user.assigned_programs.set(valid_progs)
                messages.success(request, f"Successfully updated assigned programs for Judge '{j_user.first_name or j_user.username}'!")
            return redirect('core:judge_management', institution_slug=institution.slug)

        # Action C: Save Judge Counts, Mark Entry Mode, and Assigned Judges per Program
        updated_count = 0
        program_ids = set()
        for key in request.POST.keys():
            if key.startswith('judge_count_'):
                program_ids.add(key.replace('judge_count_', ''))
            elif key.startswith('assigned_judges_'):
                program_ids.add(key.replace('assigned_judges_', ''))
            elif key.startswith('mark_entry_mode_'):
                program_ids.add(key.replace('mark_entry_mode_', ''))

        for p_id in program_ids:
            prog = Program.objects.filter(id=p_id, institution=institution).first()
            if prog:
                jc_val = request.POST.get(f'judge_count_{p_id}')
                if jc_val:
                    try:
                        jc = int(jc_val)
                        if jc > 0:
                            prog.judge_count = jc
                            prog.save(update_fields=['judge_count'])
                    except ValueError:
                        pass

                mode_val = request.POST.get(f'mark_entry_mode_{p_id}')
                if mode_val in ['OFFICIALS', 'JUDGES']:
                    prog.mark_entry_mode = mode_val
                    prog.save(update_fields=['mark_entry_mode'])

                raw_judge_ids = request.POST.getlist(f'assigned_judges_{p_id}')
                clean_judge_ids = []
                for v in raw_judge_ids:
                    if v and v.strip().isdigit():
                        val = int(v.strip())
                        if val not in clean_judge_ids:
                            clean_judge_ids.append(val)

                valid_judges = list(User.objects.filter(id__in=clean_judge_ids))
                prog.assigned_judges.set(valid_judges)

                updated_count += 1

        messages.success(request, f"Successfully updated judge counts and assignments for {updated_count} program(s)!")
        return redirect('core:judge_management', institution_slug=institution.slug)

    category_program_groups = []
    for cat in categories:
        progs = list(Program.objects.filter(category=cat, institution=institution).prefetch_related('assigned_judges').order_by('name'))
        if progs:
            category_program_groups.append({
                'category': cat,
                'programs': progs,
            })

    return render(request, 'core/judge_management.html', {
        'institution': institution,
        'competition': comp,
        'category_program_groups': category_program_groups,
        'judges': judges,
        'all_programs': all_programs,
    })


@login_required
def announce_results_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)
    
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot publish public results.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)
    
    if program.is_group:
        has_marks = GroupParticipation.objects.filter(program=program, marks__isnull=False).exists()
    else:
        has_marks = Participation.objects.filter(program=program, marks__isnull=False).exists()

    if not has_marks:
        messages.error(request, f"Cannot publish results for '{program.name}': No marks have been entered for this program yet.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    # Toggle announcement status
    program.is_announced = True
    program.announced_at = timezone.now()
    program.save()

    # Recalculate team points upon publishing
    from .services import recalculate_team_points
    recalculate_team_points(institution)

    messages.success(request, f"Results for '{program.name}' published publicly and live team leaderboard updated!")
    return redirect('core:scoring_program_list', institution_slug=institution.slug)


@login_required
def stage_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    fest_days = FestDay.objects.filter(institution=institution).order_by('day_number')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        stype = request.POST.get('stage_type', 'STAGE')
        details = request.POST.get('location_details', '')
        reserved_day_ids = request.POST.getlist('reserved_days[]')

        if name:
            stage = Stage.objects.create(
                institution=institution,
                name=name,
                stage_type=stype,
                location_details=details
            )
            if reserved_day_ids:
                days = FestDay.objects.filter(id__in=reserved_day_ids, institution=institution)
                stage.reserved_days.set(days)
            else:
                stage.reserved_days.set(fest_days)

            messages.success(request, f"Stage / Venue '{name}' created successfully!")
            return redirect('core:stage_list', institution_slug=institution.slug)

    stages = Stage.objects.filter(institution=institution).prefetch_related('reserved_days')
    return render(request, 'core/stage_list.html', {
        'institution': institution,
        'stages': stages,
        'fest_days': fest_days,
    })


@login_required
def schedule_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    schedules = ProgramSchedule.objects.filter(institution=institution).select_related('program', 'fest_day', 'stage')
    return render(request, 'core/schedule.html', {'institution': institution, 'schedules': schedules})


@login_required
def points_config_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot modify points configuration.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    config, created = PointsConfig.objects.get_or_create(institution=institution)

    if request.method == 'POST':
        enable_grades = request.POST.get('enable_grades') in ['1', 'true', 'on', True]
        config.enable_grades = enable_grades

        # Single Item Rules
        config.single_rank_1_points = int(request.POST.get('single_rank_1_points', 5))
        config.single_rank_2_points = int(request.POST.get('single_rank_2_points', 3))
        config.single_rank_3_points = int(request.POST.get('single_rank_3_points', 1))
        config.single_grade_aplus_points = int(request.POST.get('single_grade_aplus_points', 6))
        config.single_grade_a_points = int(request.POST.get('single_grade_a_points', 5))
        config.single_grade_b_points = int(request.POST.get('single_grade_b_points', 3))
        config.single_grade_c_points = int(request.POST.get('single_grade_c_points', 1))

        # Group Item Rules
        config.group_rank_1_points = int(request.POST.get('group_rank_1_points', 10))
        config.group_rank_2_points = int(request.POST.get('group_rank_2_points', 6))
        config.group_rank_3_points = int(request.POST.get('group_rank_3_points', 3))
        config.group_grade_aplus_points = int(request.POST.get('group_grade_aplus_points', 6))
        config.group_grade_a_points = int(request.POST.get('group_grade_a_points', 5))
        config.group_grade_b_points = int(request.POST.get('group_grade_b_points', 3))
        config.group_grade_c_points = int(request.POST.get('group_grade_c_points', 1))

        # Thresholds
        config.grade_aplus_threshold = int(request.POST.get('grade_aplus_threshold', 90))
        config.grade_a_threshold = int(request.POST.get('grade_a_threshold', 80))
        config.grade_b_threshold = int(request.POST.get('grade_b_threshold', 70))
        config.grade_c_threshold = int(request.POST.get('grade_c_threshold', 60))

        config.save()
        if enable_grades:
            messages.success(request, "Points & grade rules for Single and Group items updated successfully!")
        else:
            messages.success(request, "⚡ Ranks-Only Mode Active: Points awarded for 1st, 2nd & 3rd ranks only. Grade columns are hidden.")

    return render(request, 'core/points_config.html', {'institution': institution, 'config': config})


@login_required
def settings_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot access portal settings.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    competitions = Competition.objects.filter(institution=institution)
    stages = Stage.objects.filter(institution=institution)
    config, _ = PointsConfig.objects.get_or_create(institution=institution)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_points':
            enable_grades = request.POST.get('enable_grades') in ['1', 'true', 'on', True]
            config.enable_grades = enable_grades
            config.single_rank_1_points = int(request.POST.get('single_rank_1_points', 5))
            config.single_rank_2_points = int(request.POST.get('single_rank_2_points', 3))
            config.single_rank_3_points = int(request.POST.get('single_rank_3_points', 1))
            config.single_grade_aplus_points = int(request.POST.get('single_grade_aplus_points', 6))
            config.single_grade_a_points = int(request.POST.get('single_grade_a_points', 5))
            config.single_grade_b_points = int(request.POST.get('single_grade_b_points', 3))
            config.single_grade_c_points = int(request.POST.get('single_grade_c_points', 1))

            config.group_rank_1_points = int(request.POST.get('group_rank_1_points', 10))
            config.group_rank_2_points = int(request.POST.get('group_rank_2_points', 6))
            config.group_rank_3_points = int(request.POST.get('group_rank_3_points', 3))
            config.group_grade_aplus_points = int(request.POST.get('group_grade_aplus_points', 6))
            config.group_grade_a_points = int(request.POST.get('group_grade_a_points', 5))
            config.group_grade_b_points = int(request.POST.get('group_grade_b_points', 3))
            config.group_grade_c_points = int(request.POST.get('group_grade_c_points', 1))

            config.grade_aplus_threshold = int(request.POST.get('grade_aplus_threshold', 90))
            config.grade_a_threshold = int(request.POST.get('grade_a_threshold', 80))
            config.grade_b_threshold = int(request.POST.get('grade_b_threshold', 70))
            config.grade_c_threshold = int(request.POST.get('grade_c_threshold', 60))

            config.save()
            messages.success(request, "Points & grade rules for Single and Group items updated successfully!")
        elif action == 'update_chest_ranges':
            base_cats = Category.objects.filter(institution=institution, is_common=False)
            for cat in base_cats:
                val = request.POST.get(f'start_chest_no_{cat.id}')
                if val and str(val).isdigit():
                    cat.start_chest_no = int(val)
                    cat.save()
            messages.success(request, "Chest number starting ranges updated successfully!")
        elif action == 'auto_generate_chest_nos':
            from .services import auto_generate_all_chest_numbers
            count = auto_generate_all_chest_numbers(institution, overwrite=True)
            messages.success(request, f"🔄 Successfully re-generated sequential chest numbers for {count} contestants across categories!")
        elif action == 'update_participation_limits':
            comp_id = request.POST.get('competition_id')
            target_comps = Competition.objects.filter(institution=institution)
            if comp_id and str(comp_id).isdigit():
                target_comps = target_comps.filter(id=comp_id)

            max_single = int(request.POST.get('max_single_programs_per_contestant', 0) or 0)
            max_group = int(request.POST.get('max_group_programs_per_contestant', 0) or 0)
            max_total = int(request.POST.get('max_total_programs_per_contestant', 0) or 0)
            max_team_single = int(request.POST.get('max_team_participants_per_single_program', 0) or 0)
            max_team_group = int(request.POST.get('max_team_entries_per_group_program', 0) or 0)

            for comp in target_comps:
                comp.max_single_programs_per_contestant = max_single
                comp.max_group_programs_per_contestant = max_group
                comp.max_total_programs_per_contestant = max_total
                comp.max_team_participants_per_single_program = max_team_single
                comp.max_team_entries_per_group_program = max_team_group
                comp.save(update_fields=[
                    'max_single_programs_per_contestant', 
                    'max_group_programs_per_contestant', 
                    'max_total_programs_per_contestant',
                    'max_team_participants_per_single_program',
                    'max_team_entries_per_group_program'
                ])

            s_text = f"{max_single} Single Items" if max_single > 0 else "Unlimited Single Items"
            g_text = f"{max_group} Group Items" if max_group > 0 else "Unlimited Group Items"
            t_text = f", {max_total} Total Items" if max_total > 0 else ""
            ts_text = f", {max_team_single}/Team in Single" if max_team_single > 0 else ""
            tg_text = f", {max_team_group} Groups/Team in Group" if max_team_group > 0 else ""
            messages.success(request, f"🎯 Program & Team Participation Limits updated successfully ({s_text}, {g_text}{t_text}{ts_text}{tg_text})!")
        elif action == 'create_competition':
            name = request.POST.get('name')
            comp_type = request.POST.get('type')
            year = request.POST.get('year', 2026)
            max_single = int(request.POST.get('max_single_programs_per_contestant', 0) or 0)
            max_group = int(request.POST.get('max_group_programs_per_contestant', 0) or 0)
            max_total = int(request.POST.get('max_total_programs_per_contestant', 0) or 0)
            max_team_single = int(request.POST.get('max_team_participants_per_single_program', 0) or 0)
            max_team_group = int(request.POST.get('max_team_entries_per_group_program', 0) or 0)
            Competition.objects.create(
                institution=institution,
                name=name,
                type=comp_type,
                year=year,
                max_single_programs_per_contestant=max_single,
                max_group_programs_per_contestant=max_group,
                max_total_programs_per_contestant=max_total,
                max_team_participants_per_single_program=max_team_single,
                max_team_entries_per_group_program=max_team_group,
            )
        elif action == 'toggle_developer_access':
            institution.allow_developer_access = not institution.allow_developer_access
            institution.save(update_fields=['allow_developer_access'])
            if institution.allow_developer_access:
                messages.success(request, "🔓 Developer / Superadmin Support Access is now ENABLED. Developer can view your workspace to assist you.")
            else:
                messages.info(request, "🔒 Developer / Superadmin Support Access is now DISABLED.")
        elif action == 'toggle_public_suspended':
            institution.is_public_suspended = not institution.is_public_suspended
            institution.save(update_fields=['is_public_suspended'])
            if institution.is_public_suspended:
                messages.warning(request, "⏸️ Public Portal Link has been SUSPENDED. External visitors cannot access live results until enabled.")
            else:
                messages.success(request, "✅ Public Portal Link is now ACTIVE & LIVE for all visitors.")
        elif action == 'toggle_cert_addon' and (request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)):
            addon = AddOn.objects.filter(code='certificate-generation', is_active=True).first()
            if addon:
                from apps.tenants.models import GrantedAddOn, AddOnRequest
                grant, created = GrantedAddOn.objects.get_or_create(institution=institution, add_on=addon, defaults={'is_active': True})
                if not created:
                    grant.is_active = not grant.is_active
                    grant.save()
                req_obj = AddOnRequest.objects.filter(institution=institution, add_on=addon).first()
                if req_obj:
                    req_obj.status = 'approved' if grant.is_active else 'rejected'
                    req_obj.save(update_fields=['status'])
        elif action == 'toggle_contestant_addon' and (request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)):
            addon = AddOn.objects.filter(code='contestant-user-creation', is_active=True).first()
            if addon:
                from apps.tenants.models import GrantedAddOn, AddOnRequest
                grant, created = GrantedAddOn.objects.get_or_create(institution=institution, add_on=addon, defaults={'is_active': True})
                if not created:
                    grant.is_active = not grant.is_active
                    grant.save()
                req_obj = AddOnRequest.objects.filter(institution=institution, add_on=addon).first()
                if req_obj:
                    req_obj.status = 'approved' if grant.is_active else 'rejected'
                    req_obj.save(update_fields=['status'])
                status_str = "ACTIVATED (ON)" if grant.is_active else "DEACTIVATED (OFF)"
                messages.success(request, f"Contestant User Creation & Login Portal is now {status_str} for {institution.name}!")
        return redirect('core:settings', institution_slug=institution.slug)

    base_categories = Category.objects.filter(institution=institution, is_common=False).order_by('id')
    from .services import get_default_start_chest_no_for_category
    for cat in base_categories:
        cat.suggested_start_chest_no = get_default_start_chest_no_for_category(cat)

    has_cert_addon = institution.has_add_on('certificate-generation')
    cert_addon = AddOn.objects.filter(code='certificate-generation', is_active=True).first()

    has_contestant_addon = institution.has_add_on('contestant-user-creation')
    contestant_addon = AddOn.objects.filter(code='contestant-user-creation', is_active=True).first()

    context = {
        'institution': institution,
        'competitions': competitions,
        'stages': stages,
        'config': config,
        'base_categories': base_categories,
        'has_cert_addon': has_cert_addon,
        'cert_addon': cert_addon,
        'has_contestant_addon': has_contestant_addon,
        'contestant_addon': contestant_addon,
    }
    return render(request, 'core/settings.html', context)


@login_required
def settings_fests_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution).order_by('-year', '-id')
    return render(request, 'core/settings/settings_fests.html', {
        'institution': institution,
        'competitions': competitions,
    })


@login_required
def settings_participation_limits_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    competitions = Competition.objects.filter(institution=institution).order_by('-year', '-id')

    if request.method == 'POST':
        comp_id = request.POST.get('competition_id')
        target_comps = Competition.objects.filter(institution=institution)
        if comp_id and str(comp_id).isdigit():
            target_comps = target_comps.filter(id=comp_id)

        max_single = int(request.POST.get('max_single_programs_per_contestant', 0) or 0)
        max_group = int(request.POST.get('max_group_programs_per_contestant', 0) or 0)
        max_total = int(request.POST.get('max_total_programs_per_contestant', 0) or 0)
        max_team_single = int(request.POST.get('max_team_participants_per_single_program', 0) or 0)
        max_team_group = int(request.POST.get('max_team_entries_per_group_program', 0) or 0)

        for comp in target_comps:
            comp.max_single_programs_per_contestant = max_single
            comp.max_group_programs_per_contestant = max_group
            comp.max_total_programs_per_contestant = max_total
            comp.max_team_participants_per_single_program = max_team_single
            comp.max_team_entries_per_group_program = max_team_group
            comp.save(update_fields=[
                'max_single_programs_per_contestant', 
                'max_group_programs_per_contestant', 
                'max_total_programs_per_contestant',
                'max_team_participants_per_single_program',
                'max_team_entries_per_group_program'
            ])

        s_text = f"{max_single} Single Items" if max_single > 0 else "Unlimited Single Items"
        g_text = f"{max_group} Group Items" if max_group > 0 else "Unlimited Group Items"
        t_text = f", {max_total} Total Items" if max_total > 0 else ""
        ts_text = f", {max_team_single}/Team in Single" if max_team_single > 0 else ""
        tg_text = f", {max_team_group} Groups/Team in Group" if max_team_group > 0 else ""
        messages.success(request, f"🎯 Program & Team Participation Limits updated successfully ({s_text}, {g_text}{t_text}{ts_text}{tg_text})!")
        return redirect('core:settings_participation_limits', institution_slug=institution.slug)

    return render(request, 'core/settings/settings_participation_limits.html', {
        'institution': institution,
        'competitions': competitions,
    })


@login_required
def settings_chest_numbers_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_chest_ranges':
            base_cats = Category.objects.filter(institution=institution, is_common=False)
            for cat in base_cats:
                val = request.POST.get(f'start_chest_no_{cat.id}')
                if val and str(val).isdigit():
                    cat.start_chest_no = int(val)
                    cat.save()
            messages.success(request, "Chest number starting ranges updated successfully!")
        elif action == 'auto_generate_chest_nos':
            from .services import auto_generate_all_chest_numbers
            count = auto_generate_all_chest_numbers(institution, overwrite=True)
            messages.success(request, f"🔄 Successfully re-generated sequential chest numbers for {count} contestants across categories!")
        return redirect('core:settings_chest_numbers', institution_slug=institution.slug)

    base_categories = Category.objects.filter(institution=institution, is_common=False).order_by('id')
    from .services import get_default_start_chest_no_for_category
    for cat in base_categories:
        cat.suggested_start_chest_no = get_default_start_chest_no_for_category(cat)

    return render(request, 'core/settings/settings_chest_numbers.html', {
        'institution': institution,
        'base_categories': base_categories,
    })


@login_required
def settings_public_portal_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_public_suspended':
            institution.is_public_suspended = not institution.is_public_suspended
            institution.save(update_fields=['is_public_suspended'])
            if institution.is_public_suspended:
                messages.warning(request, "⏸️ Public Portal Link has been SUSPENDED. External visitors cannot access live results until enabled.")
            else:
                messages.success(request, "✅ Public Portal Link is now ACTIVE & LIVE for all visitors.")
        return redirect('core:settings_public_portal', institution_slug=institution.slug)

    return render(request, 'core/settings/settings_public_portal.html', {
        'institution': institution,
    })


@login_required
def settings_support_access_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_developer_access':
            institution.allow_developer_access = not institution.allow_developer_access
            institution.save(update_fields=['allow_developer_access'])
            if institution.allow_developer_access:
                messages.success(request, "🔓 Developer / Superadmin Support Access is now ENABLED. Developer can view your workspace to assist you.")
            else:
                messages.info(request, "🔒 Developer / Superadmin Support Access is now DISABLED.")
        return redirect('core:settings_support_access', institution_slug=institution.slug)

    return render(request, 'core/settings/settings_support_access.html', {
        'institution': institution,
    })


@login_required
def settings_addons_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_cert_addon' and (request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)):
            addon = AddOn.objects.filter(code='certificate-generation', is_active=True).first()
            if addon:
                from apps.tenants.models import GrantedAddOn, AddOnRequest
                grant, created = GrantedAddOn.objects.get_or_create(institution=institution, add_on=addon, defaults={'is_active': True})
                if not created:
                    grant.is_active = not grant.is_active
                    grant.save()
                req_obj = AddOnRequest.objects.filter(institution=institution, add_on=addon).first()
                if req_obj:
                    req_obj.status = 'approved' if grant.is_active else 'rejected'
                    req_obj.save(update_fields=['status'])
        elif action == 'toggle_contestant_addon' and (request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)):
            addon = AddOn.objects.filter(code='contestant-user-creation', is_active=True).first()
            if addon:
                from apps.tenants.models import GrantedAddOn, AddOnRequest
                grant, created = GrantedAddOn.objects.get_or_create(institution=institution, add_on=addon, defaults={'is_active': True})
                if not created:
                    grant.is_active = not grant.is_active
                    grant.save()
                req_obj = AddOnRequest.objects.filter(institution=institution, add_on=addon).first()
                if req_obj:
                    req_obj.status = 'approved' if grant.is_active else 'rejected'
                    req_obj.save(update_fields=['status'])
                status_str = "ACTIVATED (ON)" if grant.is_active else "DEACTIVATED (OFF)"
                messages.success(request, f"Contestant User Creation & Login Portal is now {status_str} for {institution.name}!")
        return redirect('core:settings_addons', institution_slug=institution.slug)

    has_cert_addon = institution.has_add_on('certificate-generation')
    cert_addon = AddOn.objects.filter(code='certificate-generation', is_active=True).first()
    has_contestant_addon = institution.has_add_on('contestant-user-creation')
    contestant_addon = AddOn.objects.filter(code='contestant-user-creation', is_active=True).first()

    return render(request, 'core/settings/settings_addons.html', {
        'institution': institution,
        'has_cert_addon': has_cert_addon,
        'cert_addon': cert_addon,
        'has_contestant_addon': has_contestant_addon,
        'contestant_addon': contestant_addon,
    })


@login_required
def settings_operations_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot access operational settings.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    competitions = Competition.objects.filter(institution=institution).order_by('-year', '-id')

    if request.method == 'POST':
        action = request.POST.get('action')
        comp_id = request.POST.get('competition_id')
        target_comps = Competition.objects.filter(institution=institution)
        if comp_id and str(comp_id).isdigit():
            target_comps = target_comps.filter(id=comp_id)

        if action == 'save_locks':
            allow_team = request.POST.get('allow_team_management') in ['1', 'true', 'on', True]
            allow_cat = request.POST.get('allow_category_management') in ['1', 'true', 'on', True]
            allow_prog = request.POST.get('allow_program_management') in ['1', 'true', 'on', True]
            allow_cont = request.POST.get('allow_contestant_registration') in ['1', 'true', 'on', True]
            allow_assign = request.POST.get('allow_program_assignment') in ['1', 'true', 'on', True]

            for comp in target_comps:
                comp.allow_team_management = allow_team
                comp.allow_category_management = allow_cat
                comp.allow_program_management = allow_prog
                comp.allow_contestant_registration = allow_cont
                comp.allow_program_assignment = allow_assign
                comp.save(update_fields=[
                    'allow_team_management',
                    'allow_category_management',
                    'allow_program_management',
                    'allow_contestant_registration',
                    'allow_program_assignment'
                ])
            messages.success(request, "🔐 Operational permissions & data entry locks updated successfully!")

        elif action == 'unlock_all':
            for comp in target_comps:
                comp.allow_team_management = True
                comp.allow_category_management = True
                comp.allow_program_management = True
                comp.allow_contestant_registration = True
                comp.allow_program_assignment = True
                comp.save(update_fields=[
                    'allow_team_management',
                    'allow_category_management',
                    'allow_program_management',
                    'allow_contestant_registration',
                    'allow_program_assignment'
                ])
            messages.success(request, "🔓 All operations (Teams, Categories, Programs, Contestants, Assignments) have been UNLOCKED.")

        elif action == 'lock_all':
            for comp in target_comps:
                comp.allow_team_management = False
                comp.allow_category_management = False
                comp.allow_program_management = False
                comp.allow_contestant_registration = False
                comp.allow_program_assignment = False
                comp.save(update_fields=[
                    'allow_team_management',
                    'allow_category_management',
                    'allow_program_management',
                    'allow_contestant_registration',
                    'allow_program_assignment'
                ])
            messages.warning(request, "🔒 All operations have been LOCKED (Read-Only / Protected Mode Active).")

        return redirect('core:settings_operations', institution_slug=institution.slug)

    active_comp = competitions.filter(is_active=True).first() or competitions.first()

    return render(request, 'core/settings/settings_operations.html', {
        'institution': institution,
        'competitions': competitions,
        'active_comp': active_comp,
    })


@login_required
def program_assign_contestants_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    eligible_cats = program.category.get_eligible_categories()

    eligible_contestants = Contestant.objects.filter(
        institution=institution,
        category__in=eligible_cats
    ).select_related('team', 'category')

    if program.is_group:
        existing_part_ids = set(GroupParticipation.objects.filter(program=program).values_list('team_id', flat=True))
    else:
        existing_part_ids = set(Participation.objects.filter(program=program).values_list('contestant_id', flat=True))

    if request.method == 'POST':
        if not institution.allows_program_assignment:
            messages.error(request, "🔒 Program assignment is currently locked in Settings.")
            return redirect('core:program_assign_contestants', institution_slug=institution.slug, program_id=program.id)

        selected_ids = request.POST.getlist('selected_ids[]')
        selected_ids_set = set(int(x) for x in selected_ids if str(x).isdigit())

        if program.is_group:
            for team_id in selected_ids_set:
                GroupParticipation.objects.get_or_create(
                    institution=institution,
                    program=program,
                    team_id=team_id
                )
            GroupParticipation.objects.filter(program=program).exclude(team_id__in=selected_ids_set).delete()
        else:
            comp = program.competition
            if comp and (comp.has_single_limit or comp.has_total_limit):
                overlimit_errors = []
                for contestant_id in selected_ids_set:
                    c = Contestant.objects.filter(id=contestant_id, institution=institution).first()
                    if not c:
                        continue
                    if contestant_id not in existing_part_ids:
                        can_add, reason = c.can_enroll_single(1)
                        if not can_add:
                            overlimit_errors.append(f"#{c.chest_no} {c.name} ({reason})")

                if overlimit_errors:
                    messages.error(
                        request,
                        f"⚠️ Participation Limit Exceeded for contestant(s): {', '.join(overlimit_errors)}. Assignment was not saved."
                    )
                    return redirect('core:program_assign_contestants', institution_slug=institution.slug, program_id=program.id)

            if program.has_team_limit:
                from collections import Counter
                selected_contestants = Contestant.objects.filter(id__in=selected_ids_set, institution=institution).select_related('team')
                team_counts = Counter(c.team for c in selected_contestants if c.team)
                team_overlimit_errors = []
                for tm, count in team_counts.items():
                    if count > program.effective_max_participants_per_team:
                        team_overlimit_errors.append(f"Team '{tm.name}' ({count} selected / Max {program.effective_max_participants_per_team} allowed)")

                if team_overlimit_errors:
                    messages.error(
                        request,
                        f"⚠️ Team Participation Limit Exceeded for program '{program.name}': {', '.join(team_overlimit_errors)}. Assignment was not saved."
                    )
                    return redirect('core:program_assign_contestants', institution_slug=institution.slug, program_id=program.id)

            for contestant_id in selected_ids_set:
                Participation.objects.get_or_create(
                    institution=institution,
                    program=program,
                    contestant_id=contestant_id
                )
            Participation.objects.filter(program=program).exclude(contestant_id__in=selected_ids_set).delete()

        messages.success(request, f"Assigned participants updated for program '{program.name}'!")
        return redirect(f"{reverse('core:assignment_hub', kwargs={'institution_slug': institution.slug})}?program_id={program.id}")

    teams = Team.objects.filter(institution=institution) if program.is_group else None

    return render(request, 'core/program_assign.html', {
        'institution': institution,
        'program': program,
        'eligible_contestants': eligible_contestants,
        'existing_part_ids': existing_part_ids,
        'teams': teams,
    })


@login_required
def contestant_assign_programs_view(request, institution_slug, contestant_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    contestant = get_object_or_404(Contestant, id=contestant_id, institution=institution)

    if request.user.is_team_leader:
        managed_team = getattr(request.user, 'managed_team', None)
        if not managed_team or contestant.team_id != managed_team.id:
            messages.error(request, "Permission Denied: You can only assign programs to your own team members.")
            return redirect('core:contestant_list', institution_slug=institution.slug)

    existing_prog_ids = set(Participation.objects.filter(contestant=contestant).values_list('program_id', flat=True))

    all_programs = Program.objects.filter(institution=institution, is_group=False).select_related('category').prefetch_related('category__included_categories')
    
    c_cat = contestant.category
    c_included_cat_ids = set(c_cat.included_categories.values_list('id', flat=True)) if c_cat.is_common else set()

    eligible_programs = []
    for prog in all_programs:
        p_cat = prog.category
        if p_cat.id == c_cat.id:
            eligible_programs.append(prog)
        elif p_cat.is_common:
            p_inc_ids = set(p_cat.included_categories.values_list('id', flat=True))
            if not p_inc_ids or c_cat.id in p_inc_ids:
                eligible_programs.append(prog)
        elif c_cat.is_common:
            if not c_included_cat_ids or p_cat.id in c_included_cat_ids:
                eligible_programs.append(prog)
        elif prog.id in existing_prog_ids:
            eligible_programs.append(prog)

    if request.method == 'POST':
        if not institution.allows_program_assignment:
            messages.error(request, "🔒 Program assignment is currently locked in Settings.")
            return redirect('core:contestant_assign_programs', institution_slug=institution.slug, contestant_id=contestant.id)

        selected_prog_ids = request.POST.getlist('selected_program_ids[]')
        selected_prog_ids_set = set(int(x) for x in selected_prog_ids if str(x).isdigit())

        comp = contestant.competition
        if comp:
            if comp.has_single_limit and len(selected_prog_ids_set) > comp.max_single_programs_per_contestant:
                messages.error(
                    request,
                    f"⚠️ Single Program Limit Exceeded: Contestant #{contestant.chest_no} ({contestant.name}) can only participate in up to {comp.max_single_programs_per_contestant} single programs (you selected {len(selected_prog_ids_set)})."
                )
                return redirect('core:contestant_assign_programs', institution_slug=institution.slug, contestant_id=contestant.id)

            if comp.has_total_limit:
                current_group_count = contestant.group_programs_count
                total_attempted = len(selected_prog_ids_set) + current_group_count
                if total_attempted > comp.max_total_programs_per_contestant:
                    messages.error(
                        request,
                        f"⚠️ Total Program Limit Exceeded: Contestant #{contestant.chest_no} ({contestant.name}) can participate in a maximum of {comp.max_total_programs_per_contestant} total programs ({len(selected_prog_ids_set)} single + {current_group_count} group = {total_attempted})."
                    )
                    return redirect('core:contestant_assign_programs', institution_slug=institution.slug, contestant_id=contestant.id)

        if contestant.team:
            team_limit_errors = []
            for prog_id in selected_prog_ids_set:
                if prog_id not in existing_prog_ids:
                    prog_obj = Program.objects.filter(id=prog_id, institution=institution).first()
                    if prog_obj and prog_obj.has_team_limit:
                        can_team, reason = prog_obj.can_team_enroll(contestant.team, additional=1)
                        if not can_team:
                            team_limit_errors.append(reason)

            if team_limit_errors:
                messages.error(
                    request,
                    f"⚠️ Team Participation Limit Exceeded: {'; '.join(team_limit_errors)}"
                )
                return redirect('core:contestant_assign_programs', institution_slug=institution.slug, contestant_id=contestant.id)

        for prog_id in selected_prog_ids_set:
            Participation.objects.get_or_create(
                institution=institution,
                program_id=prog_id,
                contestant=contestant
            )

        Participation.objects.filter(contestant=contestant).exclude(program_id__in=selected_prog_ids_set).delete()

        messages.success(request, f"Assigned programs updated for contestant #{contestant.chest_no} '{contestant.name}'!")
        return redirect('core:assignment_hub', institution_slug=institution.slug)

    return render(request, 'core/contestant_assign.html', {
        'institution': institution,
        'contestant': contestant,
        'eligible_programs': eligible_programs,
        'existing_prog_ids': existing_prog_ids,
    })


@login_required
def assignment_hub_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    programs = Program.objects.filter(institution=institution).select_related('category', 'competition')
    categories = Category.objects.filter(institution=institution).prefetch_related('included_categories')
    teams = Team.objects.filter(institution=institution)
    contestants = Contestant.objects.filter(institution=institution).select_related('team', 'category')
    
    managed_team = getattr(request.user, 'managed_team', None) if request.user.is_team_leader else None

    if managed_team:
        contestants = contestants.filter(team=managed_team)
        teams = teams.filter(id=managed_team.id)

    selected_program_id = request.GET.get('program_id')
    selected_program = None
    eligible_contestants = []
    existing_part_ids = set()
    program_teams = None

    if selected_program_id:
        selected_program = Program.objects.filter(id=selected_program_id, institution=institution).first()
        if selected_program:
            eligible_cats = selected_program.category.get_eligible_categories()
            eligible_contestants = Contestant.objects.filter(
                institution=institution,
                category__in=eligible_cats
            ).select_related('team', 'category')

            if managed_team:
                eligible_contestants = eligible_contestants.filter(team=managed_team)

            if selected_program.is_group:
                existing_part_ids = set(GroupParticipation.objects.filter(program=selected_program).values_list('team_id', flat=True))
                program_teams = teams
            else:
                existing_part_ids = set(Participation.objects.filter(program=selected_program).values_list('contestant_id', flat=True))

    return render(request, 'core/assignment_hub.html', {
        'institution': institution,
        'programs': programs,
        'categories': categories,
        'teams': teams,
        'contestants': contestants,
        'selected_program': selected_program,
        'eligible_contestants': eligible_contestants,
        'existing_part_ids': existing_part_ids,
        'program_teams': program_teams,
        'managed_team': managed_team,
    })


@login_required
def assigned_programs_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    view_mode = request.GET.get('view', 'category')

    categories = Category.objects.filter(institution=institution).prefetch_related(
        'programs', 
        'programs__single_participations__contestant',
        'programs__single_participations__contestant__team'
    )
    teams = Team.objects.filter(institution=institution).prefetch_related(
        'contestants', 
        'contestants__participations__program'
    )
    programs = Program.objects.filter(institution=institution).select_related(
        'category', 'competition'
    ).prefetch_related(
        'single_participations__contestant',
        'single_participations__contestant__team',
        'group_participations__team'
    )
    contestants = Contestant.objects.filter(institution=institution).select_related(
        'team', 'category'
    ).prefetch_related(
        'participations__program'
    )

    return render(request, 'core/assigned_programs_list.html', {
        'institution': institution,
        'view_mode': view_mode,
        'categories': categories,
        'teams': teams,
        'programs': programs,
        'contestants': contestants,
    })


def render_to_pdf(template_src, context_dict={}, filename="document.pdf", request=None):
    import io
    import re
    import urllib.parse
    from xhtml2pdf import pisa
    from django.template.loader import get_template
    from django.http import HttpResponse, HttpResponseServerError

    # Sanitize filename for Content-Disposition header
    clean_filename = re.sub(r'[\r\n\t"\\\/]', '_', str(filename)).strip()
    if not clean_filename.lower().endswith('.pdf'):
        clean_filename += '.pdf'

    encoded_filename = urllib.parse.quote(clean_filename)

    template = get_template(template_src)
    html = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        pdf_bytes = result.getvalue()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        # Allow optional inline preview via ?preview=1 or ?inline=1, otherwise default to attachment
        disposition_type = 'inline' if (request and (request.GET.get('preview') == '1' or request.GET.get('inline') == '1')) else 'attachment'
        response['Content-Disposition'] = f'{disposition_type}; filename="{clean_filename}"; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Length'] = str(len(pdf_bytes))
        return response
    return HttpResponseServerError("Error generating PDF document", content_type="text/plain")


@login_required
def download_programs_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    programs = Program.objects.filter(institution=institution).select_related('category', 'competition').order_by('category__name', 'name')
    
    context = {
        'institution': institution,
        'programs': programs,
        'generated_at': timezone.now()
    }
    filename = f"{institution.slug}_programs_list.pdf"
    return render_to_pdf('pdf/programs_pdf.html', context, filename, request=request)


@login_required
def download_contestants_teamwise_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    
    if request.user.is_team_leader:
        managed_team = getattr(request.user, 'managed_team', None)
        if managed_team:
            teams = Team.objects.filter(id=managed_team.id).prefetch_related(
                'contestants', 'contestants__category'
            )
            filename = f"{managed_team.name}_contestants.pdf"
        else:
            teams = Team.objects.none()
            filename = f"{institution.slug}_contestants.pdf"
        unassigned_contestants = []
    else:
        teams = Team.objects.filter(institution=institution).prefetch_related(
            'contestants', 'contestants__category'
        ).order_by('name')
        filename = f"{institution.slug}_contestants_teamwise.pdf"
        unassigned_contestants = Contestant.objects.filter(
            institution=institution, team__isnull=True
        ).select_related('category').order_by('chest_no')

    context = {
        'institution': institution,
        'teams': teams,
        'unassigned_contestants': unassigned_contestants,
        'generated_at': timezone.now()
    }
    return render_to_pdf('pdf/contestants_teamwise_pdf.html', context, filename, request=request)


@login_required
def download_team_results_pdf_view(request, institution_slug, team_id=None):
    institution = get_object_or_404(Institution, slug=institution_slug)
    all_teams = Team.objects.filter(institution=institution)
    
    is_team_leader = request.user.is_team_leader and hasattr(request.user, 'managed_team') and request.user.managed_team
    if is_team_leader:
        team = request.user.managed_team
    elif team_id:
        team = get_object_or_404(Team, id=team_id, institution=institution)
    else:
        team = all_teams.first()

    if not team:
        messages.error(request, "No team found.")
        return redirect('core:dashboard', institution_slug=institution.slug)

    single_parts = Participation.objects.filter(
        institution=institution,
        contestant__team=team,
        program__is_announced=True,
        marks__isnull=False
    ).select_related('program', 'program__category', 'contestant', 'contestant__team').order_by('program__category__name', 'program__name')

    group_parts = GroupParticipation.objects.filter(
        institution=institution,
        team=team,
        program__is_announced=True,
        marks__isnull=False
    ).select_related('program', 'program__category', 'captain', 'team').prefetch_related('contestants').order_by('program__category__name', 'program__name')

    detailed_rows = []
    total_announced_points = 0

    for p in single_parts:
        if p.rank or p.grade:
            pts = p.total_points
            total_announced_points += pts
            detailed_rows.append({
                'chest_no': f"#{p.contestant.chest_no}",
                'contestant_name': p.contestant.name,
                'category_name': p.program.category.name,
                'team_name': team.name,
                'item_name': p.program.name,
                'is_group': False,
                'rank': p.rank,
                'grade': p.grade,
                'points': pts
            })

    for gp in group_parts:
        if gp.rank or gp.grade:
            pts = gp.total_points
            total_announced_points += pts
            c_no = f"#{gp.captain.chest_no}" if gp.captain else "-"
            detailed_rows.append({
                'chest_no': c_no,
                'contestant_name': gp.display_name,
                'category_name': gp.program.category.name,
                'team_name': team.name,
                'item_name': gp.program.name,
                'is_group': True,
                'rank': gp.rank,
                'grade': gp.grade,
                'points': pts
            })

    from .services import get_team_standings
    standings = get_team_standings(institution, announced_only=True)
    team_position = None
    for s in standings:
        if s['team'].id == team.id:
            team_position = s['position']
            break

    context = {
        'institution': institution,
        'team': team,
        'detailed_rows': detailed_rows,
        'total_announced_points': total_announced_points,
        'team_position': team_position,
        'generated_at': timezone.now()
    }
    filename = f"{team.name}_detailed_points.pdf"
    return render_to_pdf('pdf/team_detailed_results_pdf.html', context, filename, request=request)


@login_required
def download_assigned_programs_teamwise_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    
    if request.user.is_team_leader:
        managed_team = getattr(request.user, 'managed_team', None)
        if managed_team:
            teams = Team.objects.filter(id=managed_team.id).order_by('name')
            filename = f"{managed_team.name}_assigned_programs.pdf"
        else:
            teams = Team.objects.none()
            filename = f"{institution.slug}_assigned_programs.pdf"
    else:
        teams = Team.objects.filter(institution=institution).order_by('name')
        filename = f"{institution.slug}_assigned_programs_teamwise.pdf"

    team_data = []
    for team in teams:
        categories_dict = {}
        contestants = Contestant.objects.filter(institution=institution, team=team).select_related('category').prefetch_related('participations__program').order_by('category__name', 'chest_no')
        
        for contestant in contestants:
            cat_name = contestant.category.name
            if cat_name not in categories_dict:
                categories_dict[cat_name] = []
            categories_dict[cat_name].append(contestant)

        team_data.append({
            'team': team,
            'categories': categories_dict
        })

    context = {
        'institution': institution,
        'team_data': team_data,
        'generated_at': timezone.now()
    }
    return render_to_pdf('pdf/assigned_programs_teamwise_pdf.html', context, filename, request=request)


@login_required
def download_green_room_pdf_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)
    
    if program.is_group:
        participants = GroupParticipation.objects.filter(program=program).select_related('team', 'captain')
    else:
        participants = Contestant.objects.filter(
            participations__program=program
        ).select_related('team', 'category').order_by('chest_no')

    context = {
        'institution': institution,
        'program': program,
        'participants': participants,
        'generated_at': timezone.now()
    }
    filename = f"{program.name}_green_room.pdf"
    return render_to_pdf('pdf/green_room_pdf.html', context, filename, request=request)


@login_required
def download_call_list_pdf_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if program.is_group:
        participants = GroupParticipation.objects.filter(program=program).select_related('team', 'captain')
    else:
        participants = Contestant.objects.filter(
            participations__program=program
        ).select_related('team', 'category').order_by('chest_no')

    context = {
        'institution': institution,
        'program': program,
        'participants': participants,
        'generated_at': timezone.now()
    }
    filename = f"{program.name}_call_list.pdf"
    return render_to_pdf('pdf/call_list_pdf.html', context, filename, request=request)


@login_required
def download_valuation_form_pdf_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if program.is_group:
        participants = GroupParticipation.objects.filter(program=program).select_related('team', 'captain')
    else:
        participants = Participation.objects.filter(program=program).select_related('contestant', 'contestant__team')

    context = {
        'institution': institution,
        'program': program,
        'participants': participants,
        'generated_at': timezone.now()
    }
    filename = f"{program.name}_valuation_form.pdf"
    return render_to_pdf('pdf/valuation_form_pdf.html', context, filename, request=request)


@login_required
def download_bulk_green_room_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category_id = request.GET.get('category_id')

    programs = Program.objects.filter(institution=institution).select_related('category')
    if category_id and str(category_id).isdigit():
        programs = programs.filter(category_id=int(category_id))
    programs = programs.order_by('category__name', 'name')

    # Bulk Query 1: Single participations -> contestant list
    single_parts = Participation.objects.filter(
        institution=institution
    ).select_related('contestant', 'contestant__team', 'contestant__category').order_by('contestant__chest_no')

    single_by_program = {}
    for p in single_parts:
        if p.contestant:
            single_by_program.setdefault(p.program_id, []).append(p.contestant)

    # Bulk Query 2: Group participations
    group_parts = GroupParticipation.objects.filter(
        institution=institution
    ).select_related('team', 'captain')

    group_by_program = {}
    for gp in group_parts:
        group_by_program.setdefault(gp.program_id, []).append(gp)

    programs_data = []
    for prog in programs:
        participants = group_by_program.get(prog.id, []) if prog.is_group else single_by_program.get(prog.id, [])
        programs_data.append({
            'program': prog,
            'participants': participants
        })

    context = {
        'institution': institution,
        'programs_data': programs_data,
        'generated_at': timezone.now()
    }
    filename = f"{institution.slug}_all_green_room_sheets.pdf"
    return render_to_pdf('pdf/bulk_green_room_pdf.html', context, filename, request=request)


@login_required
def download_bulk_call_list_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category_id = request.GET.get('category_id')

    programs = Program.objects.filter(institution=institution).select_related('category')
    if category_id and str(category_id).isdigit():
        programs = programs.filter(category_id=int(category_id))
    programs = programs.order_by('category__name', 'name')

    # Bulk Query 1: Single participations -> contestant list
    single_parts = Participation.objects.filter(
        institution=institution
    ).select_related('contestant', 'contestant__team', 'contestant__category').order_by('contestant__chest_no')

    single_by_program = {}
    for p in single_parts:
        if p.contestant:
            single_by_program.setdefault(p.program_id, []).append(p.contestant)

    # Bulk Query 2: Group participations
    group_parts = GroupParticipation.objects.filter(
        institution=institution
    ).select_related('team', 'captain')

    group_by_program = {}
    for gp in group_parts:
        group_by_program.setdefault(gp.program_id, []).append(gp)

    programs_data = []
    for prog in programs:
        participants = group_by_program.get(prog.id, []) if prog.is_group else single_by_program.get(prog.id, [])
        programs_data.append({
            'program': prog,
            'participants': participants
        })

    context = {
        'institution': institution,
        'programs_data': programs_data,
        'generated_at': timezone.now()
    }
    filename = f"{institution.slug}_all_call_lists.pdf"
    return render_to_pdf('pdf/bulk_call_list_pdf.html', context, filename, request=request)


@login_required
def download_bulk_valuation_form_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category_id = request.GET.get('category_id')

    programs = Program.objects.filter(institution=institution).select_related('category')
    if category_id and str(category_id).isdigit():
        programs = programs.filter(category_id=int(category_id))
    programs = programs.order_by('category__name', 'name')

    # Bulk Query 1: Single participations
    single_parts = Participation.objects.filter(
        institution=institution
    ).select_related('program', 'contestant', 'contestant__team').order_by('contestant__chest_no')

    single_by_program = {}
    for p in single_parts:
        single_by_program.setdefault(p.program_id, []).append(p)

    # Bulk Query 2: Group participations
    group_parts = GroupParticipation.objects.filter(
        institution=institution
    ).select_related('team', 'captain')

    group_by_program = {}
    for gp in group_parts:
        group_by_program.setdefault(gp.program_id, []).append(gp)

    programs_data = []
    for prog in programs:
        participants = group_by_program.get(prog.id, []) if prog.is_group else single_by_program.get(prog.id, [])
        programs_data.append({
            'program': prog,
            'participants': participants
        })

    context = {
        'institution': institution,
        'programs_data': programs_data,
        'generated_at': timezone.now()
    }
    filename = f"{institution.slug}_all_valuation_forms.pdf"
    return render_to_pdf('pdf/bulk_valuation_form_pdf.html', context, filename, request=request)


@login_required
def download_single_result_pdf_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if program.is_group:
        all_parts = list(GroupParticipation.objects.filter(
            program=program, marks__isnull=False
        ).select_related('team', 'captain').prefetch_related('contestants').order_by('rank', '-marks'))
    else:
        all_parts = list(Participation.objects.filter(
            program=program, marks__isnull=False
        ).select_related('contestant', 'contestant__team').order_by('rank', '-marks'))

    winners = [p for p in all_parts if p.rank in [1, 2, 3]]
    other_grade_holders = [p for p in all_parts if (not p.rank or p.rank > 3) and p.grade] if institution.has_grades else []

    context = {
        'institution': institution,
        'program': program,
        'winners': winners,
        'other_grade_holders': other_grade_holders,
        'generated_at': timezone.now()
    }
    filename = f"{program.name}_official_result.pdf"
    return render_to_pdf('pdf/program_result_pdf.html', context, filename, request=request)


@login_required
def download_all_results_pdf_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category_id = request.GET.get('category_id') or request.GET.get('category')

    programs = Program.objects.filter(institution=institution, is_announced=True).select_related('category')
    if category_id and str(category_id).isdigit():
        programs = programs.filter(category_id=int(category_id))
    programs = programs.order_by('category__name', 'name')

    results_data = []
    for prog in programs:
        if prog.is_group:
            all_parts = list(GroupParticipation.objects.filter(
                program=prog, marks__isnull=False
            ).select_related('team', 'captain').prefetch_related('contestants').order_by('rank', '-marks'))
        else:
            all_parts = list(Participation.objects.filter(
                program=prog, marks__isnull=False
            ).select_related('contestant', 'contestant__team').order_by('rank', '-marks'))

        winners = [p for p in all_parts if p.rank in [1, 2, 3]]
        other_grade_holders = [p for p in all_parts if (not p.rank or p.rank > 3) and p.grade] if institution.has_grades else []

        if winners or other_grade_holders:
            results_data.append({
                'program': prog,
                'winners': winners,
                'other_grade_holders': other_grade_holders
            })

    context = {
        'institution': institution,
        'results_data': results_data,
        'generated_at': timezone.now()
    }
    filename = f"{institution.slug}_all_announced_results.pdf"
    return render_to_pdf('pdf/bulk_program_results_pdf.html', context, filename, request=request)


# ---------------- Category Edit & Delete ----------------
@login_required
def category_edit_view(request, institution_slug, category_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category = get_object_or_404(Category, id=category_id, institution=institution)
    competitions = Competition.objects.filter(institution=institution)
    base_categories = Category.objects.filter(institution=institution, is_common=False).exclude(id=category.id)

    if request.method == 'POST':
        if not institution.allows_category_management:
            messages.error(request, "🔒 Category creation and editing is currently locked in Settings.")
            return redirect('core:category_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        is_common = request.POST.get('is_common') == '1'
        inc_cat_ids = request.POST.getlist('included_categories[]')

        comp = Competition.objects.filter(id=comp_id, institution=institution).first()
        if comp:
            category.competition = comp
        category.name = name
        category.description = description
        category.is_common = is_common
        category.save()

        if is_common:
            inc_cats = Category.objects.filter(id__in=inc_cat_ids, institution=institution)
            category.included_categories.set(inc_cats)
        else:
            category.included_categories.clear()

        messages.success(request, f"Category '{name}' updated successfully!")
        return redirect('core:category_list', institution_slug=institution.slug)

    return render(request, 'core/category_edit.html', {
        'institution': institution,
        'category': category,
        'competitions': competitions,
        'base_categories': base_categories
    })


@login_required
def category_delete_view(request, institution_slug, category_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category = get_object_or_404(Category, id=category_id, institution=institution)
    if not institution.allows_category_management:
        messages.error(request, "🔒 Category creation and editing is currently locked in Settings.")
        return redirect('core:category_list', institution_slug=institution.slug)

    name = category.name
    category.delete()
    messages.success(request, f"Category '{name}' deleted successfully!")
    return redirect('core:category_list', institution_slug=institution.slug)


# ---------------- Team Edit & Delete ----------------
@login_required
def team_edit_view(request, institution_slug, team_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    team = get_object_or_404(Team, id=team_id, institution=institution)
    competitions = Competition.objects.filter(institution=institution)
    team_leaders = User.objects.filter(institution=institution, role='TEAM_LEADER')

    if request.method == 'POST':
        if not institution.allows_team_management:
            messages.error(request, "🔒 Team creation and editing is currently locked in Settings.")
            return redirect('core:team_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        name = request.POST.get('name')
        code_letter = request.POST.get('code_letter', '').strip().upper()
        leader_id = request.POST.get('leader_id')

        comp = get_object_or_404(Competition, id=comp_id, institution=institution)
        team.competition = comp
        team.name = name
        team.code_letter = code_letter

        if leader_id:
            leader = User.objects.filter(id=leader_id, institution=institution, role='TEAM_LEADER').first()
            if leader:
                Team.objects.filter(institution=institution, user=leader).exclude(id=team.id).update(user=None)
                team.user = leader
            else:
                team.user = None
        else:
            team.user = None

        if request.FILES.get('logo'):
            team.logo = request.FILES['logo']

        team.save()
        messages.success(request, f"Team '{name}' updated successfully!")
        return redirect('core:team_list', institution_slug=institution.slug)

    return render(request, 'core/team_edit.html', {
        'institution': institution,
        'team': team,
        'competitions': competitions,
        'team_leaders': team_leaders
    })


@login_required
def team_delete_view(request, institution_slug, team_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    team = get_object_or_404(Team, id=team_id, institution=institution)
    if not institution.allows_team_management:
        messages.error(request, "🔒 Team creation and editing is currently locked in Settings.")
        return redirect('core:team_list', institution_slug=institution.slug)

    name = team.name
    team.delete()
    messages.success(request, f"Team '{name}' deleted successfully!")
    return redirect('core:team_list', institution_slug=institution.slug)


# ---------------- Program Edit & Delete ----------------
@login_required
def program_edit_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution)

    if request.method == 'POST':
        if not institution.allows_program_management:
            messages.error(request, "🔒 Program creation, editing, and import are currently locked in Settings.")
            return redirect('core:program_list', institution_slug=institution.slug)

        comp_id = request.POST.get('competition_id')
        cat_id = request.POST.get('category_id')
        name = request.POST.get('name')
        is_group = request.POST.get('is_group') == 'on'
        p_type = request.POST.get('program_type', 'STAGE')
        p_mode = request.POST.get('presentation_mode', 'SEQUENTIAL')
        duration = request.POST.get('duration_per_participant', 5)
        max_team = int(request.POST.get('max_participants_per_team', 0) or 0)

        comp = get_object_or_404(Competition, id=comp_id, institution=institution)
        cat = get_object_or_404(Category, id=cat_id, institution=institution)

        program.competition = comp
        program.category = cat
        program.name = name
        program.is_group = is_group
        program.program_type = p_type
        program.presentation_mode = p_mode
        program.duration_per_participant = int(duration) if str(duration).isdigit() else 5
        program.max_participants_per_team = max_team

        program.save()

        messages.success(request, f"Program '{name}' updated successfully!")
        return redirect('core:program_list', institution_slug=institution.slug)

    return render(request, 'core/program_edit.html', {
        'institution': institution,
        'program': program,
        'competitions': competitions,
        'categories': categories
    })


@login_required
def program_delete_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)
    if not institution.allows_program_management:
        messages.error(request, "🔒 Program creation, editing, and import are currently locked in Settings.")
        return redirect('core:program_list', institution_slug=institution.slug)

    name = program.name
    program.delete()
    messages.success(request, f"Program '{name}' deleted successfully!")
    return redirect('core:program_list', institution_slug=institution.slug)


# ---------------- Contestant Edit & Delete ----------------
@login_required
def contestant_edit_view(request, institution_slug, contestant_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    contestant = get_object_or_404(Contestant, id=contestant_id, institution=institution)
    competitions = Competition.objects.filter(institution=institution)
    categories = Category.objects.filter(institution=institution, is_common=False)
    teams = Team.objects.filter(institution=institution)

    if request.method == 'POST':
        if not institution.allows_contestant_registration:
            messages.error(request, "🔒 Contestant registration, editing, and bulk upload are currently locked in Settings.")
            return redirect('core:contestant_list', institution_slug=institution.slug)

        chest_no = request.POST.get('chest_no')
        name = request.POST.get('name')
        comp_id = request.POST.get('competition_id')
        team_id = request.POST.get('team_id')
        cat_id = request.POST.get('category_id')

        comp = get_object_or_404(Competition, id=comp_id, institution=institution)
        team = get_object_or_404(Team, id=team_id, institution=institution)
        cat = get_object_or_404(Category, id=cat_id, institution=institution)

        if cat.is_common:
            messages.error(request, f"Contestants cannot be assigned to Combined Category '{cat.name}'. Please choose a Base Category.")
            return redirect('core:contestant_edit', institution_slug=institution.slug, contestant_id=contestant.id)

        wa_num = request.POST.get('whatsapp_number', '').strip()

        if chest_no:
            target_no = int(chest_no)
            existing = Contestant.objects.filter(
                institution=institution,
                competition=comp,
                chest_no=target_no
            ).exclude(id=contestant.id).first()
            if existing:
                messages.error(request, f"Chest No #{target_no} is already assigned to {existing.name} ({existing.team.name}) in this festival.")
                return redirect('core:contestant_edit', institution_slug=institution.slug, contestant_id=contestant.id)
            contestant.chest_no = target_no

        contestant.name = name
        contestant.whatsapp_number = wa_num
        contestant.competition = comp
        contestant.team = team
        contestant.category = cat
        contestant.save()

        messages.success(request, f"Contestant #{contestant.chest_no} '{name}' updated successfully!")
        return redirect('core:contestant_list', institution_slug=institution.slug)

    return render(request, 'core/contestant_edit.html', {
        'institution': institution,
        'contestant': contestant,
        'competitions': competitions,
        'categories': categories,
        'teams': teams
    })


@login_required
def contestant_delete_view(request, institution_slug, contestant_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    contestant = get_object_or_404(Contestant, id=contestant_id, institution=institution)
    if not institution.allows_contestant_registration:
        messages.error(request, "🔒 Contestant registration, editing, and bulk upload are currently locked in Settings.")
        return redirect('core:contestant_list', institution_slug=institution.slug)

    c_name = contestant.name
    chest_no = contestant.chest_no
    contestant.delete()
    messages.success(request, f"Contestant #{chest_no} '{c_name}' deleted successfully!")
    return redirect('core:contestant_list', institution_slug=institution.slug)


# ==============================================================================
# RESULTS, TEAM STANDINGS, TOPPERS & SCORE BALANCER AI
# ==============================================================================

@login_required
def program_results_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    categories = Category.objects.filter(institution=institution)
    
    category_id = request.GET.get('category')
    view_mode = request.GET.get('view', 'announced')
    
    programs = Program.objects.filter(institution=institution).select_related('category')
    if category_id:
        programs = programs.filter(category_id=category_id)
    if view_mode == 'announced':
        programs = programs.filter(is_announced=True)

    program_results = []
    for prog in programs:
        if prog.is_group:
            participations = GroupParticipation.objects.filter(
                program=prog, marks__isnull=False
            ).select_related('team').prefetch_related('contestants').order_by('rank', '-marks')
        else:
            participations = Participation.objects.filter(
                program=prog, marks__isnull=False
            ).select_related('contestant', 'contestant__team').order_by('rank', '-marks')

        if participations.exists():
            program_results.append({
                'program': prog,
                'participations': participations
            })

    return render(request, 'core/program_results.html', {
        'institution': institution,
        'categories': categories,
        'selected_category_id': category_id,
        'view_mode': view_mode,
        'program_results': program_results
    })


@login_required
def team_standings_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    view_mode = request.GET.get('view', 'announced')
    announced_only = (view_mode == 'announced')

    n_param = request.GET.get('n_results') or request.GET.get('n')
    limit_n = None
    if n_param and n_param.isdigit():
        limit_n = int(n_param)

    from .services import get_team_standings
    team_data = get_team_standings(institution, announced_only=announced_only, limit_n_results=limit_n)

    total_announced_programs = Program.objects.filter(institution=institution, is_announced=True).count()
    total_completed_programs = Program.objects.filter(institution=institution).filter(
        Q(single_participations__marks__isnull=False) | Q(group_participations__marks__isnull=False)
    ).distinct().count()
    max_results = total_announced_programs if announced_only else total_completed_programs

    n_presets = [opt for opt in [5, 10, 15, 20, 25, 30, 40, 50] if opt < max_results]

    return render(request, 'core/team_standings.html', {
        'institution': institution,
        'teams': team_data,
        'top_three': team_data[:3] if len(team_data) >= 3 else team_data,
        'view_mode': view_mode,
        'selected_n': limit_n,
        'max_results': max_results,
        'total_announced_programs': total_announced_programs,
        'total_completed_programs': total_completed_programs,
        'n_presets': n_presets,
    })


@login_required
def team_points_cards_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    n_param = request.GET.get('n_results') or request.GET.get('n')
    limit_n = None
    if n_param and n_param.isdigit():
        limit_n = int(n_param)

    from .services import get_team_standings
    team_data = get_team_standings(institution, announced_only=True, limit_n_results=limit_n)

    total_announced_programs = Program.objects.filter(institution=institution, is_announced=True).count()

    comp = Competition.objects.filter(institution=institution, is_active=True).first()
    fest_title = comp.name if comp else institution.name

    n_presets = [opt for opt in [5, 10, 15, 20, 25, 30, 40, 50] if opt < total_announced_programs]

    return render(request, 'core/team_points_cards.html', {
        'institution': institution,
        'teams': team_data,
        'top_three': team_data[:3] if len(team_data) >= 3 else team_data,
        'remaining_teams': team_data[3:] if len(team_data) > 3 else [],
        'selected_n': limit_n,
        'total_announced_programs': total_announced_programs,
        'fest_title': fest_title,
        'n_presets': n_presets,
    })


@login_required
def toppers_list_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    categories = Category.objects.filter(institution=institution)
    category_id = request.GET.get('category')
    stage_type = request.GET.get('stage_type')
    view_mode = request.GET.get('view', 'announced')
    announced_only = (view_mode == 'announced')

    contestants_qs = Contestant.objects.filter(institution=institution).select_related('team', 'category')
    if category_id:
        contestants_qs = contestants_qs.filter(category_id=category_id)

    toppers_data = []
    for c in contestants_qs:
        parts = Participation.objects.filter(contestant=c, marks__isnull=False).select_related('program')
        if announced_only:
            parts = parts.filter(program__is_announced=True)
        if stage_type in ['STAGE', 'OFF_STAGE']:
            parts = parts.filter(program__program_type=stage_type)

        tot_pts = 0
        r1 = r2 = r3 = 0
        for p in parts:
            if p.rank == 1: r1 += 1
            elif p.rank == 2: r2 += 1
            elif p.rank == 3: r3 += 1
            tot_pts += p.total_points

        if tot_pts > 0 or r1 > 0 or r2 > 0 or r3 > 0:
            toppers_data.append({
                'contestant': c,
                'points': tot_pts,
                'r1': r1,
                'r2': r2,
                'r3': r3
            })

    toppers_data.sort(key=lambda x: (x['points'], x['r1'], x['r2'], x['r3']), reverse=True)

    overall_champion = toppers_data[0] if toppers_data else None

    cat_champions = []
    for cat in categories:
        cat_contestants = [t for t in toppers_data if t['contestant'].category.id == cat.id]
        if cat_contestants:
            cat_champions.append({
                'category': cat,
                'champion': cat_contestants[0]
            })

    return render(request, 'core/toppers_list.html', {
        'institution': institution,
        'categories': categories,
        'selected_category_id': category_id,
        'selected_stage_type': stage_type,
        'toppers': toppers_data,
        'overall_champion': overall_champion,
        'cat_champions': cat_champions,
        'view_mode': view_mode
    })


@login_required
def manage_announcements_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot access the Public Announcement Hub.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    if request.method == 'POST' and request.POST.get('action') == 'toggle_public_suspended':
        institution.is_public_suspended = not institution.is_public_suspended
        institution.save(update_fields=['is_public_suspended'])
        if institution.is_public_suspended:
            messages.warning(request, "⏸️ Public Portal Link has been SUSPENDED. External visitors cannot access live results until enabled.")
        else:
            messages.success(request, "✅ Public Portal Link is now ACTIVE & LIVE for all visitors.")
        return redirect('core:manage_announcements', institution_slug=institution.slug)

    programs = Program.objects.filter(institution=institution).select_related('category').order_by('category__name', 'name')

    for p in programs:
        if p.is_group:
            p.has_marks = GroupParticipation.objects.filter(program=p, marks__isnull=False).exists()
            p.marked_count = GroupParticipation.objects.filter(program=p, marks__isnull=False).count()
        else:
            p.has_marks = Participation.objects.filter(program=p, marks__isnull=False).exists()
            p.marked_count = Participation.objects.filter(program=p, marks__isnull=False).count()

    announced_count = programs.filter(is_announced=True).count()
    total_programs = programs.count()

    suggested_announcements = get_top_5_balancing_announcement_suggestions(institution)

    return render(request, 'core/manage_announcements.html', {
        'institution': institution,
        'programs': programs,
        'announced_count': announced_count,
        'total_programs': total_programs,
        'suggested_announcements': suggested_announcements
    })


@login_required
def toggle_program_announcement_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot publish public results.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    if not program.is_announced:
        if program.is_group:
            has_marks = GroupParticipation.objects.filter(program=program, marks__isnull=False).exists()
        else:
            has_marks = Participation.objects.filter(program=program, marks__isnull=False).exists()

        if not has_marks:
            messages.error(request, f"Cannot publish results for '{program.name}': No marks have been entered for this program yet.")
            next_url = request.META.get('HTTP_REFERER') or redirect('core:manage_announcements', institution_slug=institution.slug)
            return redirect(next_url)

    program.is_announced = not program.is_announced
    if program.is_announced:
        program.announced_at = timezone.now()
        if not program.result_number:
            from django.db.models import Max
            max_num = Program.objects.filter(
                institution=institution,
                competition=program.competition,
                is_announced=True
            ).aggregate(Max('result_number'))['result_number__max'] or 0
            program.result_number = max_num + 1
        messages.success(request, f"📢 Results for '{program.name}' are now PUBLICLY ANNOUNCED (Result #{program.result_number})!")
    else:
        messages.info(request, f"🔒 Results for '{program.name}' are now hidden from public view.")
    program.save()

    from .services import recalculate_team_points
    recalculate_team_points(institution)

    next_url = request.META.get('HTTP_REFERER') or redirect('core:manage_announcements', institution_slug=institution.slug)
    return redirect(next_url)


@login_required
def update_program_result_number_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    raw_val = request.POST.get('result_number', '').strip()
    if not raw_val or not raw_val.isdigit():
        return JsonResponse({'success': False, 'message': 'Please enter a valid positive number.'}, status=400)

    new_number = int(raw_val)
    if new_number <= 0:
        return JsonResponse({'success': False, 'message': 'Result number must be greater than 0.'}, status=400)

    # Check if another program in this festival already has this result number
    existing = Program.objects.filter(
        institution=institution,
        competition=program.competition,
        result_number=new_number
    ).exclude(id=program.id).first()

    if existing:
        return JsonResponse({
            'success': False,
            'message': f"That number is already fixed for result {existing.name}"
        }, status=400)

    program.result_number = new_number
    program.save(update_fields=['result_number'])

    return JsonResponse({
        'success': True,
        'result_number': program.result_number,
        'message': f"Result number for '{program.name}' successfully updated to #{program.result_number}."
    })


def get_top_5_balancing_announcement_suggestions(institution):
    """
    Calculates top 5 unannounced completed programs that best balance 
    the current public team scores and create maximum suspense on the public leaderboard.
    """
    teams = list(Team.objects.filter(institution=institution))
    if not teams:
        return []

    public_team_scores = {}
    for t in teams:
        pts = 0
        for c in Contestant.objects.filter(institution=institution, team=t):
            parts = Participation.objects.filter(contestant=c, marks__isnull=False, program__is_announced=True)
            pts += sum(p.total_points for p in parts if p.rank or p.grade)
        for gp in GroupParticipation.objects.filter(team=t, marks__isnull=False, program__is_announced=True):
            pts += gp.total_points
        public_team_scores[t.id] = pts

    unannounced_programs = Program.objects.filter(
        institution=institution, is_announced=False
    ).filter(
        Q(single_participations__marks__isnull=False) | Q(group_participations__marks__isnull=False)
    ).distinct()

    suggestions = []
    for prog in unannounced_programs:
        simulated_gains = {t.id: 0 for t in teams}
        
        if prog.is_group:
            gps = GroupParticipation.objects.filter(program=prog, marks__isnull=False)
            for gp in gps:
                if gp.team_id and (gp.rank or gp.grade):
                    simulated_gains[gp.team_id] = simulated_gains.get(gp.team_id, 0) + gp.total_points
        else:
            ps = Participation.objects.filter(program=prog, marks__isnull=False).select_related('contestant')
            for p in ps:
                if p.contestant and p.contestant.team_id and (p.rank or p.grade):
                    simulated_gains[p.contestant.team_id] = simulated_gains.get(p.contestant.team_id, 0) + p.total_points

        simulated_scores = {t.id: public_team_scores[t.id] + simulated_gains[t.id] for t in teams}
        sorted_scores = sorted(simulated_scores.values(), reverse=True)
        
        if len(sorted_scores) >= 2:
            gap_1st_2nd = sorted_scores[0] - sorted_scores[1]
            gap_1st_3rd = sorted_scores[0] - sorted_scores[2] if len(sorted_scores) >= 3 else gap_1st_2nd
            balance_score = -(gap_1st_2nd * 1.5 + gap_1st_3rd * 0.5)
        else:
            balance_score = 0

        total_prog_pts = sum(simulated_gains.values())
        final_priority = balance_score + (total_prog_pts * 0.1)

        impact_items = []
        for t in teams:
            gain = simulated_gains.get(t.id, 0)
            if gain > 0:
                impact_items.append(f"{t.name}: +{gain} pts")

        suggestions.append({
            'program': prog,
            'total_pts': total_prog_pts,
            'priority': round(final_priority, 2),
            'impact_summary': ", ".join(impact_items) if impact_items else "No team points awarded",
            'top_gap_after': sorted_scores[0] - sorted_scores[1] if len(sorted_scores) >= 2 else 0
        })

    suggestions.sort(key=lambda x: x['priority'], reverse=True)
    return suggestions[:5]


@login_required
def announcement_balancer_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.user.is_judge:
        messages.error(request, "Permission Denied: Judges cannot access the Score Balancer AI.")
        return redirect('core:scoring_program_list', institution_slug=institution.slug)

    suggested_announcements = get_top_5_balancing_announcement_suggestions(institution)

    public_team_scores = get_team_standings(institution, announced_only=True)

    return render(request, 'core/announcement_balancer.html', {
        'institution': institution,
        'suggested_announcements': suggested_announcements,
        'public_team_scores': public_team_scores
    })


@login_required
def shareable_results_view(request, institution_slug):
    from .models import CustomResultTemplate

    institution = get_object_or_404(Institution, slug=institution_slug)
    comp = Competition.objects.filter(institution=institution, is_active=True).first() or Competition.objects.filter(institution=institution).first()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'upload_custom_template' and request.FILES.get('template_image'):
            if comp:
                current_count = CustomResultTemplate.objects.filter(competition=comp).count()
                if current_count >= 10:
                    messages.error(request, "Maximum limit of 10 custom poster templates reached. Please delete an existing template to add a new one.")
                else:
                    template_name = request.POST.get('template_name', '').strip() or f"Custom Poster {current_count + 1}"
                    CustomResultTemplate.objects.create(
                        institution=institution,
                        competition=comp,
                        name=template_name,
                        image=request.FILES['template_image']
                    )
                    messages.success(request, f"Custom Template '{template_name}' uploaded successfully!")
            return redirect('core:shareable_results', institution_slug=institution.slug)

        elif action == 'delete_custom_template':
            template_id = request.POST.get('template_id')
            tmpl = CustomResultTemplate.objects.filter(id=template_id, competition=comp).first()
            if tmpl:
                t_name = tmpl.name
                tmpl.image.delete(save=False)
                tmpl.delete()
                messages.success(request, f"Custom Template '{t_name}' deleted successfully!")
            return redirect('core:shareable_results', institution_slug=institution.slug)

    custom_templates = list(CustomResultTemplate.objects.filter(competition=comp)) if comp else []

    announced_progs = list(Program.objects.filter(institution=institution, is_announced=True).select_related('category', 'competition').distinct().order_by('result_number', 'announced_at', 'id'))
    
    # Auto-assign result_number if any program is missing one
    assigned_nums = set(p.result_number for p in announced_progs if p.result_number)
    next_num = 1
    for p in announced_progs:
        if not p.result_number:
            while next_num in assigned_nums:
                next_num += 1
            p.result_number = next_num
            p.save(update_fields=['result_number'])
            assigned_nums.add(next_num)

    programs = sorted(announced_progs, key=lambda p: (p.result_number or 999999, p.announced_at or timezone.now()))

    cards_data = []
    for prog in programs:
        if prog.is_group:
            winners = GroupParticipation.objects.filter(
                program=prog,
                rank__in=[1, 2, 3]
            ).select_related('team').prefetch_related('contestants').order_by('rank')
            
            winners_list = []
            for gp in winners:
                members = ", ".join([c.name for c in gp.contestants.all()])
                winners_list.append({
                    'rank': gp.rank,
                    'name': gp.display_name,
                    'team': gp.team.name if gp.team else '',
                    'members': members,
                    'marks': gp.marks,
                    'grade': gp.grade
                })
        else:
            winners = Participation.objects.filter(
                program=prog,
                rank__in=[1, 2, 3]
            ).select_related('contestant', 'contestant__team').order_by('rank')
            
            winners_list = []
            for p in winners:
                winners_list.append({
                    'rank': p.rank,
                    'name': p.contestant.name if p.contestant else '',
                    'team': p.contestant.team.name if p.contestant and p.contestant.team else '',
                    'chest_no': p.contestant.chest_no if p.contestant else '',
                    'members': '',
                    'marks': p.marks,
                    'grade': p.grade
                })

        if winners_list:
            cards_data.append({
                'program': prog,
                'winners': winners_list
            })

    return render(request, 'core/shareable_results.html', {
        'institution': institution,
        'cards_data': cards_data,
        'competition': comp,
        'custom_templates': custom_templates
    })


# ==============================================================================
# WINNER CERTIFICATE GENERATION STUDIO (PAID ADD-ON)
# ==============================================================================

def _get_certificate_winners_data(institution, competition, config, category_id=None, program_id=None, place=None, specific_part_id=None, is_group=None):
    # Allowed ranks filter
    if place:
        try:
            allowed_ranks = [int(place)]
        except (ValueError, TypeError):
            allowed_ranks = [1, 2, 3]
    else:
        try:
            allowed_ranks = [int(r.strip()) for r in config.include_places.split(',') if r.strip().isdigit()]
        except Exception:
            allowed_ranks = [1, 2, 3]

    announced_progs_qs = Program.objects.filter(institution=institution, is_announced=True).select_related('category', 'competition')
    if category_id:
        announced_progs_qs = announced_progs_qs.filter(category_id=category_id)
    if program_id:
        announced_progs_qs = announced_progs_qs.filter(id=program_id)
    
    programs = list(announced_progs_qs.order_by('category__name', 'name'))
    
    certificates = []
    cert_index = 1
    
    place_meta = {
        1: {
            'rank_display': '1st Place',
            'place_title': '1ST PLACE WINNER',
            'place_suffix': 'ST',
            'color_theme': 'gold',
            'color_hex': '#d97706',
            'color_accent': '#f59e0b',
            'color_dark': '#78350f',
            'color_light': '#fef3c7',
            'color_badge_bg': '#fffbeb',
            'gradient': 'linear-gradient(135deg, #b45309, #f59e0b, #d97706)',
            'ribbon_bg': '#d97706',
            'trophy_color': '#f59e0b',
        },
        2: {
            'rank_display': '2nd Place',
            'place_title': '2ND PLACE WINNER',
            'place_suffix': 'ND',
            'color_theme': 'silver',
            'color_hex': '#475569',
            'color_accent': '#94a3b8',
            'color_dark': '#1e293b',
            'color_light': '#f1f5f9',
            'color_badge_bg': '#f8fafc',
            'gradient': 'linear-gradient(135deg, #334155, #94a3b8, #64748b)',
            'ribbon_bg': '#64748b',
            'trophy_color': '#94a3b8',
        },
        3: {
            'rank_display': '3rd Place',
            'place_title': '3RD PLACE WINNER',
            'place_suffix': 'RD',
            'color_theme': 'bronze',
            'color_hex': '#9a3412',
            'color_accent': '#d97706',
            'color_dark': '#7c2d12',
            'color_light': '#ffedd5',
            'color_badge_bg': '#fff7ed',
            'gradient': 'linear-gradient(135deg, #7c2d12, #ea580c, #9a3412)',
            'ribbon_bg': '#c2410c',
            'trophy_color': '#cd7f32',
        },
    }

    issue_date_str = config.issue_date.strftime('%d %B %Y') if config.issue_date else timezone.now().strftime('%d %B %Y')
    fest_name = competition.name if competition else institution.name
    fest_year = str(competition.year) if competition else str(timezone.now().year)

    for prog in programs:
        if prog.is_group:
            gp_qs = GroupParticipation.objects.filter(
                program=prog,
                rank__in=allowed_ranks
            ).select_related('team').prefetch_related('contestants').order_by('rank', '-marks')
            
            if specific_part_id and is_group:
                gp_qs = gp_qs.filter(id=specific_part_id)

            for gp in gp_qs:
                pm = place_meta.get(gp.rank, place_meta[1])
                grade_str = f"Grade {gp.grade}" if gp.grade else "Qualified"
                
                para = config.paragraph_template or "In recognition of outstanding performance and securing {rank_display} ({grade_display}) in {program_name} ({category_name}) at {institution_name} during {fest_name} {fest_year}."
                try:
                    citation = para.format(
                        rank_display=pm['rank_display'],
                        grade_display=grade_str,
                        program_name=prog.name,
                        category_name=prog.category.name,
                        institution_name=institution.name,
                        fest_name=fest_name,
                        fest_year=fest_year
                    )
                except Exception:
                    citation = f"In recognition of outstanding performance and securing {pm['rank_display']} ({grade_str}) in {prog.name} ({prog.category.name}) at {institution.name} during {fest_name} {fest_year}."

                members_list = [c.name for c in gp.contestants.all()]
                members_str = ", ".join(members_list) if members_list else ""

                certificates.append({
                    'index': cert_index,
                    'is_group': True,
                    'part_id': gp.id,
                    'rank': gp.rank,
                    'rank_display': pm['rank_display'],
                    'place_title': pm['place_title'],
                    'place_suffix': pm['place_suffix'],
                    'color_theme': pm['color_theme'],
                    'color_hex': pm['color_hex'],
                    'color_accent': pm['color_accent'],
                    'color_dark': pm['color_dark'],
                    'color_light': pm['color_light'],
                    'color_badge_bg': pm['color_badge_bg'],
                    'gradient': pm['gradient'],
                    'ribbon_bg': pm['ribbon_bg'],
                    'trophy_color': pm['trophy_color'],
                    'recipient_name': gp.display_name,
                    'members_str': members_str,
                    'chest_no': f"Team {gp.team.code_letter}" if (gp.team and gp.team.code_letter) else "",
                    'team_name': gp.team.name if gp.team else '',
                    'program_name': prog.name,
                    'program_id': prog.id,
                    'category_name': prog.category.name,
                    'category_id': prog.category.id,
                    'grade': gp.grade or '',
                    'grade_display': grade_str,
                    'marks': gp.marks,
                    'citation_text': citation,
                    'issue_date_str': issue_date_str,
                })
                cert_index += 1
        else:
            p_qs = Participation.objects.filter(
                program=prog,
                rank__in=allowed_ranks
            ).select_related('contestant', 'contestant__team', 'contestant__category').order_by('rank', '-marks')

            if specific_part_id and not is_group:
                p_qs = p_qs.filter(id=specific_part_id)

            for p in p_qs:
                pm = place_meta.get(p.rank, place_meta[1])
                grade_str = f"Grade {p.grade}" if p.grade else "Qualified"
                
                para = config.paragraph_template or "In recognition of outstanding performance and securing {rank_display} ({grade_display}) in {program_name} ({category_name}) at {institution_name} during {fest_name} {fest_year}."
                try:
                    citation = para.format(
                        rank_display=pm['rank_display'],
                        grade_display=grade_str,
                        program_name=prog.name,
                        category_name=prog.category.name,
                        institution_name=institution.name,
                        fest_name=fest_name,
                        fest_year=fest_year
                    )
                except Exception:
                    citation = f"In recognition of outstanding performance and securing {pm['rank_display']} ({grade_str}) in {prog.name} ({prog.category.name}) at {institution.name} during {fest_name} {fest_year}."

                certificates.append({
                    'index': cert_index,
                    'is_group': False,
                    'part_id': p.id,
                    'rank': p.rank,
                    'rank_display': pm['rank_display'],
                    'place_title': pm['place_title'],
                    'place_suffix': pm['place_suffix'],
                    'color_theme': pm['color_theme'],
                    'color_hex': pm['color_hex'],
                    'color_accent': pm['color_accent'],
                    'color_dark': pm['color_dark'],
                    'color_light': pm['color_light'],
                    'color_badge_bg': pm['color_badge_bg'],
                    'gradient': pm['gradient'],
                    'ribbon_bg': pm['ribbon_bg'],
                    'trophy_color': pm['trophy_color'],
                    'recipient_name': p.contestant.name,
                    'members_str': '',
                    'chest_no': f"#{p.contestant.chest_no}",
                    'team_name': p.contestant.team.name if p.contestant.team else '',
                    'program_name': prog.name,
                    'program_id': prog.id,
                    'category_name': prog.category.name,
                    'category_id': prog.category.id,
                    'grade': p.grade or '',
                    'grade_display': grade_str,
                    'marks': p.marks,
                    'citation_text': citation,
                    'issue_date_str': issue_date_str,
                })
                cert_index += 1

    return certificates


@login_required
def certificate_studio_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    comp = Competition.objects.filter(institution=institution, is_active=True).first() or Competition.objects.filter(institution=institution).first()
    has_cert_addon = institution.has_add_on('certificate-generation') or request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)
    cert_addon = AddOn.objects.filter(code='certificate-generation', is_active=True).first()

    # Get or create certificate config
    config, _ = CertificateConfig.objects.get_or_create(
        institution=institution,
        defaults={'competition': comp}
    )
    if comp and config.competition != comp:
        config.competition = comp
        config.save(update_fields=['competition'])

    # Handle POST Actions
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'request_cert_addon':
            if cert_addon:
                from apps.tenants.models import AddOnRequest
                req_obj, created = AddOnRequest.objects.get_or_create(
                    institution=institution,
                    add_on=cert_addon,
                    defaults={
                        'status': 'pending',
                        'contact_person': request.user.get_full_name() or request.user.username,
                        'contact_phone': getattr(institution, 'phone', '') or getattr(request.user, 'phone', '') or '',
                    }
                )
                if not created and req_obj.status == 'rejected':
                    req_obj.status = 'pending'
                    req_obj.save(update_fields=['status'])
            messages.success(request, "🎉 Your purchase / activation request for 'Winner Certificate Generation Studio' (PRO) has been registered! The developer will review and activate it.")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        if not has_cert_addon and action in ['update_settings', 'update_config', 'upload_custom_template', 'upload_signatures']:
            messages.warning(request, "🔒 Winner Certificate Generation Studio is a PRO Add-On. Please unlock the add-on to save custom settings.")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        if action in ['update_settings', 'update_config']:
            config.title = request.POST.get('title', config.title).strip()
            config.subtitle = request.POST.get('subtitle', config.subtitle).strip()
            config.paragraph_template = request.POST.get('paragraph_template', config.paragraph_template).strip()
            config.signatory_1_title = request.POST.get('signatory_1_title', config.signatory_1_title).strip()
            config.signatory_1_name = request.POST.get('signatory_1_name', config.signatory_1_name).strip()
            config.signatory_2_title = request.POST.get('signatory_2_title', config.signatory_2_title).strip()
            config.signatory_2_name = request.POST.get('signatory_2_name', config.signatory_2_name).strip()
            config.mode = request.POST.get('mode', config.mode)

            issue_date_raw = request.POST.get('issue_date')
            if issue_date_raw:
                try:
                    config.issue_date = datetime.strptime(issue_date_raw, '%Y-%m-%d').date()
                except Exception:
                    pass

            offset_raw = request.POST.get('custom_text_offset_top')
            if offset_raw and str(offset_raw).isdigit():
                config.custom_text_offset_top = int(offset_raw)

            places = request.POST.getlist('include_places')
            if places:
                config.include_places = ",".join(places)

            config.save()
            messages.success(request, "Certificate settings updated successfully!")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        elif action == 'upload_custom_template':
            if request.FILES.get('template_image'):
                config.template_image = request.FILES['template_image']
                config.mode = 'custom'
                config.save()
                messages.success(request, "Custom certificate background template uploaded successfully!")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        elif action == 'remove_custom_template':
            if config.template_image:
                config.template_image.delete(save=False)
                config.template_image = None
                config.mode = 'code'
                config.save()
                messages.success(request, "Custom template removed. Switched back to Built-in Code Design.")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        elif action == 'upload_signatures':
            if request.FILES.get('signatory_1_signature'):
                config.signatory_1_signature = request.FILES['signatory_1_signature']
            if request.FILES.get('signatory_2_signature'):
                config.signatory_2_signature = request.FILES['signatory_2_signature']
            config.save()
            messages.success(request, "Signatures uploaded successfully!")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        elif action == 'remove_signature_1':
            if config.signatory_1_signature:
                config.signatory_1_signature.delete(save=False)
                config.signatory_1_signature = None
                config.save()
                messages.success(request, "Signatory 1 signature cleared.")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

        elif action == 'remove_signature_2':
            if config.signatory_2_signature:
                config.signatory_2_signature.delete(save=False)
                config.signatory_2_signature = None
                config.save()
                messages.success(request, "Signatory 2 signature cleared.")
            return redirect('core:certificate_studio', institution_slug=institution.slug)

    # Filtering parameters
    category_id = request.GET.get('category')
    program_id = request.GET.get('program')
    place_filter = request.GET.get('place')

    categories = Category.objects.filter(institution=institution).order_by('name')
    programs_qs = Program.objects.filter(institution=institution, is_announced=True).select_related('category')
    if category_id and str(category_id).isdigit():
        programs_qs = programs_qs.filter(category_id=category_id)
    programs = list(programs_qs.order_by('name'))

    # Retrieve all certificates matching filters
    certificates = _get_certificate_winners_data(
        institution=institution,
        competition=comp,
        config=config,
        category_id=int(category_id) if category_id and str(category_id).isdigit() else None,
        program_id=int(program_id) if program_id and str(program_id).isdigit() else None,
        place=int(place_filter) if place_filter and str(place_filter).isdigit() else None
    )

    # Sample preview certificates for 1st, 2nd, and 3rd place
    sample_preview_1 = next((c for c in certificates if c['rank'] == 1), None)
    sample_preview_2 = next((c for c in certificates if c['rank'] == 2), None)
    sample_preview_3 = next((c for c in certificates if c['rank'] == 3), None)

    # Fallback preview mocks if no announced winners yet
    sample_issue_date = config.issue_date.strftime('%d %B %Y') if config.issue_date else timezone.now().strftime('%d %B %Y')
    fest_name = comp.name if comp else institution.name
    fest_year = str(comp.year) if comp else str(timezone.now().year)

    if not sample_preview_1:
        sample_preview_1 = {
            'index': 1, 'is_group': False, 'rank': 1, 'rank_display': '1st Place', 'place_title': '1ST PLACE WINNER',
            'place_suffix': 'ST', 'color_theme': 'gold', 'color_hex': '#d97706', 'color_accent': '#f59e0b',
            'color_dark': '#78350f', 'color_light': '#fef3c7', 'color_badge_bg': '#fffbeb',
            'gradient': 'linear-gradient(135deg, #b45309, #f59e0b, #d97706)', 'ribbon_bg': '#d97706',
            'trophy_color': '#f59e0b', 'recipient_name': 'Muhammed Nihal K', 'members_str': '',
            'chest_no': '#1012', 'team_name': 'Red House', 'program_name': 'Quran Recitation (Hafs)',
            'category_name': 'Junior Category', 'grade': 'A+', 'grade_display': 'Grade A+', 'marks': 96,
            'citation_text': f"In recognition of outstanding performance and securing 1st Place (Grade A+) in Quran Recitation (Hafs) (Junior Category) at {institution.name} during {fest_name} {fest_year}.",
            'issue_date_str': sample_issue_date
        }

    if not sample_preview_2:
        sample_preview_2 = {
            'index': 2, 'is_group': False, 'rank': 2, 'rank_display': '2nd Place', 'place_title': '2ND PLACE WINNER',
            'place_suffix': 'ND', 'color_theme': 'silver', 'color_hex': '#475569', 'color_accent': '#94a3b8',
            'color_dark': '#1e293b', 'color_light': '#f1f5f9', 'color_badge_bg': '#f8fafc',
            'gradient': 'linear-gradient(135deg, #334155, #94a3b8, #64748b)', 'ribbon_bg': '#64748b',
            'trophy_color': '#94a3b8', 'recipient_name': 'Ahmad Rayan P', 'members_str': '',
            'chest_no': '#1018', 'team_name': 'Blue House', 'program_name': 'Elocution English',
            'category_name': 'Junior Category', 'grade': 'A', 'grade_display': 'Grade A', 'marks': 89,
            'citation_text': f"In recognition of outstanding performance and securing 2nd Place (Grade A) in Elocution English (Junior Category) at {institution.name} during {fest_name} {fest_year}.",
            'issue_date_str': sample_issue_date
        }

    if not sample_preview_3:
        sample_preview_3 = {
            'index': 3, 'is_group': False, 'rank': 3, 'rank_display': '3rd Place', 'place_title': '3RD PLACE WINNER',
            'place_suffix': 'RD', 'color_theme': 'bronze', 'color_hex': '#9a3412', 'color_accent': '#d97706',
            'color_dark': '#7c2d12', 'color_light': '#ffedd5', 'color_badge_bg': '#fff7ed',
            'gradient': 'linear-gradient(135deg, #7c2d12, #ea580c, #9a3412)', 'ribbon_bg': '#c2410c',
            'trophy_color': '#cd7f32', 'recipient_name': 'Zainul Abid', 'members_str': '',
            'chest_no': '#1025', 'team_name': 'Green House', 'program_name': 'Pencil Drawing',
            'category_name': 'Senior Category', 'grade': 'B', 'grade_display': 'Grade B', 'marks': 78,
            'citation_text': f"In recognition of outstanding performance and securing 3rd Place (Grade B) in Pencil Drawing (Senior Category) at {institution.name} during {fest_name} {fest_year}.",
            'issue_date_str': sample_issue_date
        }

    context = {
        'institution': institution,
        'competition': comp,
        'config': config,
        'has_cert_addon': has_cert_addon,
        'cert_addon': cert_addon,
        'categories': categories,
        'programs': programs,
        'certificates': certificates,
        'total_certificates': len(certificates),
        'sample_preview_1': sample_preview_1,
        'sample_preview_2': sample_preview_2,
        'sample_preview_3': sample_preview_3,
        'selected_category_id': int(category_id) if category_id and str(category_id).isdigit() else '',
        'selected_program_id': int(program_id) if program_id and str(program_id).isdigit() else '',
        'selected_place': int(place_filter) if place_filter and str(place_filter).isdigit() else '',
    }
    return render(request, 'core/certificate_studio.html', context)


@login_required
def download_certificate_pdf_view(request, institution_slug):
    # Pure vector print engine is used for certificates; redirect directly to print view
    query_str = request.META.get('QUERY_STRING', '')
    url = reverse('core:print_certificates', kwargs={'institution_slug': institution_slug})
    if query_str:
        url += f"?{query_str}"
    return redirect(url)


@login_required
def print_certificates_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    comp = Competition.objects.filter(institution=institution, is_active=True).first() or Competition.objects.filter(institution=institution).first()
    config, _ = CertificateConfig.objects.get_or_create(institution=institution, defaults={'competition': comp})

    has_cert_addon = institution.has_add_on('certificate-generation') or request.user.is_superuser or (hasattr(request.user, 'is_developer') and request.user.is_developer)
    if not has_cert_addon:
        messages.warning(request, "🔒 Winner Certificate Studio is a PRO Add-On. Please purchase or unlock the feature to print certificates.")
        return redirect('core:certificate_studio', institution_slug=institution.slug)

    category_id = request.GET.get('category')
    program_id = request.GET.get('program')
    place = request.GET.get('place')
    part_id = request.GET.get('part_id')
    is_group = request.GET.get('is_group') == '1'
    autoprint = request.GET.get('autoprint', '1') == '1'

    certificates = _get_certificate_winners_data(
        institution=institution,
        competition=comp,
        config=config,
        category_id=int(category_id) if category_id and str(category_id).isdigit() else None,
        program_id=int(program_id) if program_id and str(program_id).isdigit() else None,
        place=int(place) if place and str(place).isdigit() else None,
        specific_part_id=int(part_id) if part_id and str(part_id).isdigit() else None,
        is_group=is_group if part_id else None
    )

    if not certificates:
        messages.warning(request, "No announced winners found matching the selected certificate filter.")
        return redirect('core:certificate_studio', institution_slug=institution.slug)

    context = {
        'institution': institution,
        'competition': comp,
        'config': config,
        'certificates': certificates,
        'autoprint': autoprint,
    }
    return render(request, 'core/certificate_print_view.html', context)




# ==============================================================================
# FEST SCHEDULE & AUTO-SCHEDULER ENGINE
# ==============================================================================

from .schedule_utils import (
    get_program_assigned_count,
    calculate_program_duration,
    detect_all_clashes,
    generate_smart_auto_schedule
)

@login_required
def manage_schedule_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)

    fest_days = FestDay.objects.filter(institution=institution).order_by('day_number')
    stages = Stage.objects.filter(institution=institution).prefetch_related('reserved_days').order_by('stage_type', 'name')
    programs = Program.objects.filter(institution=institution).select_related('category', 'schedule', 'schedule__fest_day', 'schedule__stage').all()

    # Pre-cache participant counts to eliminate N+1 queries
    part_counts = dict(
        Participation.objects.filter(institution=institution)
        .values('program_id')
        .annotate(c=Count('id'))
        .values_list('program_id', 'c')
    )
    group_counts = dict(
        GroupParticipation.objects.filter(institution=institution)
        .values('program_id')
        .annotate(c=Count('id'))
        .values_list('program_id', 'c')
    )

    program_list = []
    scheduled_count = 0
    for p in programs:
        assigned_count = group_counts.get(p.id, 0) if p.is_group else part_counts.get(p.id, 0)
        
        # Calculate duration in-memory
        if p.presentation_mode == 'SIMULTANEOUS':
            calc_dur = p.duration_per_participant + (p.buffer_margin_minutes or 0)
        else:
            cnt = assigned_count if assigned_count > 0 else 1
            calc_dur = (cnt * p.duration_per_participant) + (p.buffer_margin_minutes or 0)
        calc_dur = max(calc_dur, 5)

        has_sched = hasattr(p, 'schedule') and p.schedule is not None
        if has_sched:
            scheduled_count += 1

        program_list.append({
            'program': p,
            'assigned_count': assigned_count,
            'calculated_duration': calc_dur,
            'has_schedule': has_sched,
            'schedule': p.schedule if has_sched else None
        })

    clash_data = detect_all_clashes(institution)

    # Bulk fetch timetable schedules to prevent N+1 query loops
    all_schedules = list(ProgramSchedule.objects.filter(institution=institution).select_related('program', 'program__category'))
    schedule_map = {}
    for s in all_schedules:
        schedule_map.setdefault((s.fest_day_id, s.stage_id), []).append(s)

    timetable_by_day = []
    next_times_map = {}
    base_date = datetime.today().date()

    for day in fest_days:
        day_stages = []
        day_default_str = day.start_time.strftime("%H:%M")
        for stage in stages:
            schedules = sorted(schedule_map.get((day.id, stage.id), []), key=lambda x: x.start_time)
            day_stages.append({
                'stage': stage,
                'schedules': schedules
            })

            key = f"{day.id}_{stage.id}"
            if schedules:
                latest_end = max(s.end_time for s in schedules)
                next_dt = datetime.combine(base_date, latest_end) + timedelta(minutes=1)
                next_times_map[key] = next_dt.time().strftime("%H:%M")
            else:
                next_times_map[key] = day_default_str

        timetable_by_day.append({
            'day': day,
            'stages': day_stages
        })

    next_times_json = json.dumps(next_times_map)

    return render(request, 'core/manage_schedule.html', {
        'institution': institution,
        'fest_days': fest_days,
        'stages': stages,
        'program_list': program_list,
        'total_programs': len(programs),
        'scheduled_count': scheduled_count,
        'clash_data': clash_data,
        'timetable_by_day': timetable_by_day,
        'next_times_json': next_times_json,
    })


@login_required
def add_fest_day_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        day_number = request.POST.get('day_number')
        date_str = request.POST.get('date')
        name = request.POST.get('name', '').strip()
        start_time_str = request.POST.get('start_time', '09:00')
        end_time_str = request.POST.get('end_time', '21:00')

        if day_number:
            parsed_date = None
            if date_str:
                try:
                    parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            try:
                st_time = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                st_time = time(9, 0)

            try:
                en_time = datetime.strptime(end_time_str, '%H:%M').time()
            except ValueError:
                en_time = time(21, 0)

            comp = Competition.objects.filter(institution=institution, is_active=True).first() or Competition.objects.filter(institution=institution).first()
            if not comp:
                comp = Competition.objects.create(institution=institution, name="Main Fest", type="ON", year=2026)

            FestDay.objects.get_or_create(
                institution=institution,
                competition=comp,
                day_number=int(day_number),
                defaults={'date': parsed_date, 'name': name, 'start_time': st_time, 'end_time': en_time}
            )
            messages.success(request, f"Fest Day #{day_number} added successfully!")

    return redirect('core:manage_schedule', institution_slug=institution.slug)


@login_required
def fest_day_edit_view(request, institution_slug, day_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    day = get_object_or_404(FestDay, id=day_id, institution=institution)

    if request.method == 'POST':
        day_number = request.POST.get('day_number')
        date_str = request.POST.get('date')
        name = request.POST.get('name', '').strip()
        start_time_str = request.POST.get('start_time', '09:00')
        end_time_str = request.POST.get('end_time', '21:00')

        if day_number:
            parsed_date = None
            if date_str:
                try:
                    parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            try:
                st_time = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                st_time = day.start_time

            try:
                en_time = datetime.strptime(end_time_str, '%H:%M').time()
            except ValueError:
                en_time = day.end_time

            day.day_number = int(day_number)
            day.date = parsed_date
            day.name = name
            day.start_time = st_time
            day.end_time = en_time
            day.save()

            messages.success(request, f"Fest Day #{day.day_number} updated successfully!")
            return redirect('core:manage_schedule', institution_slug=institution.slug)

    return render(request, 'core/fest_day_edit.html', {
        'institution': institution,
        'day': day,
    })


@login_required
def fest_day_edit_view(request, institution_slug, day_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    day = get_object_or_404(FestDay, id=day_id, institution=institution)

    if request.method == 'POST':
        day_number = request.POST.get('day_number')
        date_str = request.POST.get('date')
        name = request.POST.get('name', '').strip()
        start_time_str = request.POST.get('start_time', '09:00')
        end_time_str = request.POST.get('end_time', '21:00')

        if day_number:
            parsed_date = None
            if date_str:
                try:
                    parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            try:
                st_time = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                st_time = day.start_time

            try:
                en_time = datetime.strptime(end_time_str, '%H:%M').time()
            except ValueError:
                en_time = day.end_time

            day.day_number = int(day_number)
            day.date = parsed_date
            day.name = name
            day.start_time = st_time
            day.end_time = en_time
            day.save()

            messages.success(request, f"Fest Day #{day.day_number} updated successfully!")
            return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=setup-tab")

    return render(request, 'core/fest_day_edit.html', {
        'institution': institution,
        'day': day,
    })


@login_required
def delete_fest_day_view(request, institution_slug, day_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    day = get_object_or_404(FestDay, id=day_id, institution=institution)
    day_num = day.day_number
    day.delete()
    messages.success(request, f"Fest Day #{day_num} deleted successfully.")
    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=setup-tab")


@login_required
def add_stage_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    fest_days = FestDay.objects.filter(institution=institution).order_by('day_number')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        stage_type = request.POST.get('stage_type', 'STAGE')
        location_details = request.POST.get('location_details', '').strip()
        reserved_day_ids = request.POST.getlist('reserved_days[]')

        if name:
            stage = Stage.objects.create(
                institution=institution,
                name=name,
                stage_type=stage_type,
                location_details=location_details
            )
            if reserved_day_ids:
                days = FestDay.objects.filter(id__in=reserved_day_ids, institution=institution)
                stage.reserved_days.set(days)
            else:
                stage.reserved_days.set(fest_days)

            messages.success(request, f"Venue / Stage '{name}' ({stage_type}) added successfully!")

    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=setup-tab")


@login_required
def stage_edit_view(request, institution_slug, stage_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    stage = get_object_or_404(Stage, id=stage_id, institution=institution)
    fest_days = FestDay.objects.filter(institution=institution).order_by('day_number')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        stage_type = request.POST.get('stage_type', 'STAGE')
        location_details = request.POST.get('location_details', '').strip()
        reserved_day_ids = request.POST.getlist('reserved_days[]')

        if name:
            stage.name = name
            stage.stage_type = stage_type
            stage.location_details = location_details
            stage.save()

            if reserved_day_ids:
                days = FestDay.objects.filter(id__in=reserved_day_ids, institution=institution)
                stage.reserved_days.set(days)
            else:
                stage.reserved_days.clear()

            messages.success(request, f"Stage / Venue '{stage.name}' updated successfully!")
            return redirect('core:stage_list', institution_slug=institution.slug)

    reserved_day_ids = set(stage.reserved_days.values_list('id', flat=True))
    return render(request, 'core/stage_edit.html', {
        'institution': institution,
        'stage': stage,
        'fest_days': fest_days,
        'reserved_day_ids': reserved_day_ids,
    })


@login_required
def delete_stage_view(request, institution_slug, stage_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    stage = get_object_or_404(Stage, id=stage_id, institution=institution)
    st_name = stage.name
    stage.delete()
    messages.success(request, f"Venue '{st_name}' deleted successfully.")
    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=setup-tab")


@login_required
def save_program_schedule_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        program_id = request.POST.get('program_id')
        fest_day_id = request.POST.get('fest_day_id')
        stage_id = request.POST.get('stage_id')
        start_time_str = request.POST.get('start_time')

        if program_id and fest_day_id and stage_id and start_time_str:
            program = get_object_or_404(Program, id=program_id, institution=institution)
            fest_day = get_object_or_404(FestDay, id=fest_day_id, institution=institution)
            stage = get_object_or_404(Stage, id=stage_id, institution=institution)

            try:
                start_t = datetime.strptime(start_time_str, '%H:%M').time()
            except ValueError:
                messages.error(request, "Invalid start time format.")
                return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=manual-tab")

            calc_mins = calculate_program_duration(program)
            end_t = (datetime.combine(datetime.today(), start_t) + timedelta(minutes=calc_mins)).time()

            ProgramSchedule.objects.update_or_create(
                institution=institution,
                program=program,
                defaults={
                    'fest_day': fest_day,
                    'stage': stage,
                    'start_time': start_t,
                    'end_time': end_t,
                    'total_duration_minutes': calc_mins
                }
            )
            messages.success(request, f"Schedule saved for '{program.name}'!")

    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=manual-tab")


@login_required
def delete_program_schedule_view(request, institution_slug, schedule_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    sched = get_object_or_404(ProgramSchedule, id=schedule_id, institution=institution)
    prog_name = sched.program.name
    sched.delete()
    messages.success(request, f"Schedule for '{prog_name}' removed.")
    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=manual-tab")


@login_required
def run_auto_scheduler_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        res = generate_smart_auto_schedule(institution)
        if 'error' in res:
            messages.error(request, res['error'])
        else:
            sched_count = res.get('scheduled_count', 0)
            skip_count = res.get('skipped_count', 0)
            messages.success(request, f"🤖 Smart Auto-Scheduler completed! Successfully scheduled {sched_count} programs.")
            if skip_count > 0:
                messages.warning(request, f"Could not fit {skip_count} programs into available time slots. Consider adding another fest day or stage.")

    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=timetable-tab")


@login_required
def clear_all_schedules_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if request.method == 'POST':
        count = ProgramSchedule.objects.filter(institution=institution).count()
        ProgramSchedule.objects.filter(institution=institution).delete()
        messages.success(request, f"Cleared all {count} program schedules.")

    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=timetable-tab")


@login_required
def update_program_duration_view(request, institution_slug, program_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    program = get_object_or_404(Program, id=program_id, institution=institution)

    if request.method == 'POST':
        program_type = request.POST.get('program_type', 'STAGE')
        presentation_mode = request.POST.get('presentation_mode', 'SEQUENTIAL')
        dur_per_part = request.POST.get('duration_per_participant', '5')
        buffer_mins = request.POST.get('buffer_margin_minutes', '0')
        preferred_stage_id = request.POST.get('preferred_stage_id', '')

        program.program_type = program_type
        program.presentation_mode = presentation_mode
        program.duration_per_participant = max(int(dur_per_part), 1) if str(dur_per_part).isdigit() else 5
        program.buffer_margin_minutes = max(int(buffer_mins), 0) if str(buffer_mins).isdigit() else 0

        if preferred_stage_id:
            program.preferred_stage_id = int(preferred_stage_id)
        else:
            program.preferred_stage = None

        program.save()

        calc_dur = calculate_program_duration(program)
        if hasattr(program, 'schedule') and program.schedule is not None:
            sched = program.schedule
            sched.total_duration_minutes = calc_dur
            s_dt = datetime.combine(datetime.today(), sched.start_time)
            sched.end_time = (s_dt + timedelta(minutes=calc_dur)).time()
            sched.save()

        messages.success(request, f"Schedule settings for '{program.name}' updated.")

    return redirect(f"{reverse('core:manage_schedule', kwargs={'institution_slug': institution.slug})}?tab=manual-tab")


from django.urls import reverse

@login_required
def group_assign_view(request, institution_slug, program_id=None):
    institution = get_object_or_404(Institution, slug=institution_slug)
    group_programs = Program.objects.filter(institution=institution, is_group=True).select_related('category', 'competition')
    
    managed_team = getattr(request.user, 'managed_team', None) if request.user.is_team_leader else None

    if managed_team:
        teams = Team.objects.filter(id=managed_team.id)
    else:
        teams = Team.objects.filter(institution=institution)
    
    req_program_id = request.GET.get('program_id') or program_id
    selected_program = None
    if req_program_id:
        selected_program = Program.objects.filter(id=req_program_id, institution=institution, is_group=True).first()
    if not selected_program and group_programs.exists():
        selected_program = group_programs.first()

    selected_team_id = request.GET.get('team_id')
    if managed_team:
        selected_team = managed_team
    else:
        selected_team = Team.objects.filter(id=selected_team_id, institution=institution).first() if selected_team_id else teams.first()

    eligible_contestants = []
    if selected_program:
        eligible_cats = selected_program.category.get_eligible_categories()
        eligible_contestants = Contestant.objects.filter(
            institution=institution,
            category__in=eligible_cats
        ).select_related('team', 'category')
        if selected_team:
            eligible_contestants = eligible_contestants.filter(team=selected_team)

    existing_group_participations = []
    if selected_program:
        existing_group_participations = GroupParticipation.objects.filter(
            program=selected_program
        ).select_related('team', 'captain').prefetch_related('contestants')
        if managed_team:
            existing_group_participations = existing_group_participations.filter(team=managed_team)

    if request.method == 'POST':
        action = request.POST.get('action', 'save_group')
        if action == 'save_group':
            prog_id = request.POST.get('program_id')
            team_id = request.POST.get('team_id')
            group_name = request.POST.get('group_name', '').strip()
            captain_id = request.POST.get('captain_id')
            member_ids = request.POST.getlist('member_ids[]')
            group_part_id = request.POST.get('group_part_id')

            prog = get_object_or_404(Program, id=prog_id, institution=institution)
            if managed_team:
                tm = managed_team
            else:
                tm = get_object_or_404(Team, id=team_id, institution=institution)

            capt = Contestant.objects.filter(id=captain_id, institution=institution, team=tm).first() if captain_id else None

            if group_part_id:
                gp = get_object_or_404(GroupParticipation, id=group_part_id, institution=institution)
                if managed_team and gp.team_id != managed_team.id:
                    messages.error(request, "Permission Denied: You cannot modify group entries for other teams.")
                    return redirect('core:group_assign', institution_slug=institution.slug)
            else:
                if prog.has_team_limit:
                    current_team_groups = GroupParticipation.objects.filter(program=prog, team=tm).count()
                    if current_team_groups + 1 > prog.effective_max_participants_per_team:
                        messages.error(request, f"⚠️ Team Entry Limit Exceeded: Team '{tm.name}' has reached the maximum allowed ({prog.effective_max_participants_per_team}) entries for program '{prog.name}'.")
                        return redirect(f"{reverse('core:group_assign', kwargs={'institution_slug': institution.slug})}?program_id={prog.id}&team_id={tm.id}")
                gp = GroupParticipation(institution=institution, program=prog, team=tm)

            comp = prog.competition
            all_candidate_ids = set(int(x) for x in member_ids if str(x).isdigit())
            if capt:
                all_candidate_ids.add(capt.id)

            if comp and (comp.has_group_limit or comp.has_total_limit):
                overlimit_errors = []
                existing_gp_members = set(gp.contestants.values_list('id', flat=True)) if (group_part_id and gp.id) else set()
                if group_part_id and gp.id and gp.captain_id:
                    existing_gp_members.add(gp.captain_id)

                for cid in all_candidate_ids:
                    if cid not in existing_gp_members:
                        c_obj = Contestant.objects.filter(id=cid, institution=institution).first()
                        if c_obj:
                            can_add, reason = c_obj.can_enroll_group(1)
                            if not can_add:
                                overlimit_errors.append(f"#{c_obj.chest_no} {c_obj.name} ({reason})")

                if overlimit_errors:
                    messages.error(request, f"⚠️ Group Program Limit Exceeded for contestant(s): {', '.join(overlimit_errors)}. Group entry was not saved.")
                    return redirect(f"{reverse('core:group_assign', kwargs={'institution_slug': institution.slug})}?program_id={prog.id}&team_id={tm.id}")

            gp.group_name = group_name
            gp.captain = capt
            gp.save()

            if member_ids:
                m_contestants = Contestant.objects.filter(id__in=member_ids, institution=institution, team=tm)
                gp.contestants.set(m_contestants)

            if capt and not gp.contestants.filter(id=capt.id).exists():
                gp.contestants.add(capt)

            messages.success(request, f"Group entry '{gp.display_name}' saved successfully!")
            return redirect(f"{reverse('core:group_assign', kwargs={'institution_slug': institution.slug})}?program_id={prog.id}&team_id={tm.id}")

    return render(request, 'core/group_assign.html', {
        'institution': institution,
        'group_programs': group_programs,
        'selected_program': selected_program,
        'teams': teams,
        'selected_team': selected_team,
        'eligible_contestants': eligible_contestants,
        'existing_group_participations': existing_group_participations,
    })


@login_required
def delete_group_participation_view(request, institution_slug, group_part_id):
    institution = get_object_or_404(Institution, slug=institution_slug)
    gp = get_object_or_404(GroupParticipation, id=group_part_id, institution=institution)
    prog_id = gp.program.id
    disp_name = gp.display_name
    gp.delete()
    messages.success(request, f"Group entry '{disp_name}' deleted.")
    return redirect(f"{reverse('core:group_assign', kwargs={'institution_slug': institution.slug})}?program_id={prog_id}")


from django.http import JsonResponse

@login_required
def api_get_next_chest_no_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    category_id = request.GET.get('category_id')
    if category_id:
        category = Category.objects.filter(id=category_id, institution=institution, is_common=False).first()
        if category:
            from .services import get_next_chest_number
            next_no = get_next_chest_number(category)
            return JsonResponse({'next_chest_no': next_no})
    return JsonResponse({'next_chest_no': ''})


@login_required
def team_results_view(request, institution_slug, team_id=None):
    institution = get_object_or_404(Institution, slug=institution_slug)
    all_teams = Team.objects.filter(institution=institution)
    
    is_team_leader = request.user.is_team_leader and hasattr(request.user, 'managed_team') and request.user.managed_team
    if is_team_leader:
        team = request.user.managed_team
    elif team_id:
        team = get_object_or_404(Team, id=team_id, institution=institution)
    else:
        team = all_teams.first()

    if not team:
        messages.error(request, "No team found.")
        return redirect('core:dashboard', institution_slug=institution.slug)

    # 1. Single Participations (Announced Only)
    single_parts = Participation.objects.filter(
        institution=institution,
        contestant__team=team,
        program__is_announced=True,
        marks__isnull=False
    ).select_related('program', 'program__category', 'contestant', 'contestant__team').order_by('program__category__name', 'program__name')

    # 2. Group Participations (Announced Only)
    group_parts = GroupParticipation.objects.filter(
        institution=institution,
        team=team,
        program__is_announced=True,
        marks__isnull=False
    ).select_related('program', 'program__category', 'captain', 'team').prefetch_related('contestants').order_by('program__category__name', 'program__name')

    detailed_rows = []
    total_announced_points = 0

    for p in single_parts:
        if p.rank or p.grade:
            pts = p.total_points
            total_announced_points += pts
            detailed_rows.append({
                'chest_no': f"#{p.contestant.chest_no}",
                'contestant_name': p.contestant.name,
                'category_name': p.program.category.name,
                'team_name': team.name,
                'item_name': p.program.name,
                'is_group': False,
                'rank': p.rank,
                'grade': p.grade,
                'points': pts
            })

    for gp in group_parts:
        if gp.rank or gp.grade:
            pts = gp.total_points
            total_announced_points += pts
            c_no = f"#{gp.captain.chest_no}" if gp.captain else "-"
            detailed_rows.append({
                'chest_no': c_no,
                'contestant_name': gp.display_name,
                'category_name': gp.program.category.name,
                'team_name': team.name,
                'item_name': gp.program.name,
                'is_group': True,
                'rank': gp.rank,
                'grade': gp.grade,
                'points': pts
            })

    from .services import get_team_standings
    standings = get_team_standings(institution, announced_only=True)
    team_position = None
    for s in standings:
        if s['team'].id == team.id:
            team_position = s['position']
            break

    return render(request, 'core/team_results.html', {
        'institution': institution,
        'team': team,
        'all_teams': all_teams,
        'is_team_leader': is_team_leader,
        'detailed_rows': detailed_rows,
        'total_announced_points': total_announced_points,
        'team_position': team_position,
        'standings': standings,
    })


@login_required
def help_guide_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    return render(request, 'core/help_guide.html', {'institution': institution})


@login_required
def generate_contestant_credentials_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    if not (request.user.is_institution_admin or request.user.is_developer):
        messages.error(request, "Access Denied: Institution Admin clearance required.")
        return redirect('core:contestant_list', institution_slug=institution.slug)

    from django.utils.text import slugify
    from apps.tenants.models import AddOn, AddOnRequest

    has_contestant_addon = institution.has_add_on('contestant-user-creation')
    contestant_addon = AddOn.objects.filter(code='contestant-user-creation', is_active=True).first()
    addon_request = AddOnRequest.objects.filter(institution=institution, add_on__code='contestant-user-creation').first() if contestant_addon else None

    unregistered_contestants = Contestant.objects.filter(
        institution=institution,
        user_account__isnull=True
    ).select_related('team', 'category').order_by('chest_no')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'request_contestant_addon':
            if contestant_addon:
                contact_p = request.POST.get('contact_person') or getattr(institution, 'contact_name_val', 'Admin')
                contact_ph = request.POST.get('contact_phone') or institution.phone
                notes = request.POST.get('notes', '')
                AddOnRequest.objects.update_or_create(
                    institution=institution,
                    add_on=contestant_addon,
                    defaults={
                        'status': 'pending',
                        'contact_person': contact_p,
                        'contact_phone': contact_ph,
                        'notes': notes,
                    }
                )
                messages.success(request, f"🎉 Instant activation request for '{contestant_addon.name}' submitted successfully! Our developer will review and activate it shortly.")
            return redirect('core:generate_contestant_credentials', institution_slug=institution.slug)

        if not has_contestant_addon:
            messages.error(request, "👑 PRO Add-On Required: Please unlock the Contestant User Creation & Personal Portal feature to generate login credentials.")
            return redirect('core:generate_contestant_credentials', institution_slug=institution.slug)

        generated_count = 0
        for c in unregistered_contestants:
            # Format Name as Username (clean & slugified)
            clean_name = slugify(c.name).replace('-', '_')
            if not clean_name:
                clean_name = f"contestant_{c.chest_no}"

            username = clean_name
            # If username is taken, append chest number
            if User.objects.filter(username=username).exists():
                username = f"{clean_name}_{c.chest_no}"

            # Password is Chest Number
            pwd = str(c.chest_no)

            user = User.objects.create_user(
                username=username,
                password=pwd,
                role='CONTESTANT',
                institution=institution,
                contestant=c,
                is_approved=True,
                email=f"chest{c.chest_no}@{institution.slug}.local"
            )
            generated_count += 1

        if generated_count > 0:
            messages.success(request, f"🎉 Successfully generated login accounts for {generated_count} contestants!")
        else:
            messages.info(request, "All contestants already have active login accounts.")

        return redirect('core:generate_contestant_credentials', institution_slug=institution.slug)

    # Fetch all contestants with active credentials to display on the page
    registered_contestants = Contestant.objects.filter(
        institution=institution,
        user_account__isnull=False
    ).select_related('team', 'category', 'user_account').order_by('chest_no')

    import urllib.parse
    host = request.get_host()
    scheme = 'https' if request.is_secure() else 'http'
    login_url = f"{scheme}://{host}/auth/login/"

    credentials_list = []
    for c in registered_contestants:
        wa_link = None
        if c.whatsapp_number:
            clean_phone = "".join(filter(str.isdigit, str(c.whatsapp_number)))
            if clean_phone:
                if len(clean_phone) == 10:
                    clean_phone = "91" + clean_phone
                message = (
                    f"Assalamu Alaikum / Greetings *{c.name}*,\n\n"
                    f"Here are your login credentials for *{institution.name}*:\n\n"
                    f"👤 *Username:* `{c.user_account.username}`\n"
                    f"🔑 *Password:* `{c.chest_no}`\n"
                    f"🏷️ *Chest No:* #{c.chest_no}\n"
                    f"🏫 *Team:* {c.team.name}\n\n"
                    f"🌐 *Login Portal:* {login_url}\n\n"
                    f"Log in to view your enrolled programs, stage venues, timings, and official results!"
                )
                encoded_msg = urllib.parse.quote(message)
                wa_link = f"https://api.whatsapp.com/send?phone={clean_phone}&text={encoded_msg}"

        credentials_list.append({
            'contestant': c,
            'username': c.user_account.username,
            'password': str(c.chest_no),
            'whatsapp_phone': c.whatsapp_number,
            'whatsapp_link': wa_link
        })

    return render(request, 'core/contestant_credentials_report.html', {
        'institution': institution,
        'unregistered_count': unregistered_contestants.count(),
        'credentials_list': credentials_list,
        'has_contestant_addon': has_contestant_addon,
        'contestant_addon': contestant_addon,
        'addon_request': addon_request,
    })


@login_required
def contestant_personal_dashboard_view(request, institution_slug):
    institution = get_object_or_404(Institution, slug=institution_slug)
    
    # Resolve contestant (either logged in contestant or admin viewing specific contestant)
    if request.user.is_contestant and request.user.contestant:
        contestant = request.user.contestant
    else:
        contestant_id = request.GET.get('contestant_id')
        if contestant_id:
            contestant = Contestant.objects.filter(id=contestant_id, institution=institution).first()
        else:
            contestant = getattr(request.user, 'contestant', None)

    if not contestant:
        messages.error(request, "No contestant profile linked to your user account.")
        return redirect('landing_page')

    # 1. Gather Enrolled Programs & Timings
    single_parts = contestant.participations.select_related(
        'program', 'program__category', 'program__schedule', 
        'program__schedule__stage', 'program__schedule__fest_day'
    )
    group_parts = contestant.group_entries.select_related(
        'program', 'program__category', 'program__schedule', 
        'program__schedule__stage', 'program__schedule__fest_day'
    )

    enrolled_programs = []
    
    for p in single_parts:
        sched = getattr(p.program, 'schedule', None)
        enrolled_programs.append({
            'program': p.program,
            'format': 'Single Event',
            'is_group': False,
            'type': p.program.get_program_type_display(),
            'schedule': sched,
            'code_letter': p.code_letter,
            'result': p if p.program.is_announced else None
        })

    for gp in group_parts:
        sched = getattr(gp.program, 'schedule', None)
        enrolled_programs.append({
            'program': gp.program,
            'format': 'Group Event',
            'is_group': True,
            'type': gp.program.get_program_type_display(),
            'schedule': sched,
            'code_letter': gp.code_letter,
            'result': gp if gp.program.is_announced else None
        })

    # 2. Gather Published Results
    published_results = []
    total_individual_points = 0
    ranks_count = {1: 0, 2: 0, 3: 0}
    grades_count = {'A+': 0, 'A': 0, 'B': 0, 'C': 0}

    for item in enrolled_programs:
        res = item['result']
        if res and (res.rank or res.grade):
            pts = res.total_points
            if not item['is_group']:
                total_individual_points += pts

            if res.rank in ranks_count:
                ranks_count[res.rank] += 1
            if res.grade in grades_count:
                grades_count[res.grade] += 1

            published_results.append({
                'program_name': item['program'].name,
                'category_name': item['program'].category.name,
                'format': item['format'],
                'code_letter': res.code_letter,
                'marks': res.marks,
                'rank': res.rank,
                'grade': res.grade,
                'points': pts,
            })

    return render(request, 'core/contestant_personal_dashboard.html', {
        'institution': institution,
        'contestant': contestant,
        'enrolled_programs': enrolled_programs,
        'published_results': published_results,
        'total_individual_points': total_individual_points,
        'ranks_count': ranks_count,
        'aplus_count': grades_count.get('A+', 0),
        'a_count': grades_count.get('A', 0),
        'b_count': grades_count.get('B', 0),
        'c_count': grades_count.get('C', 0),
    })


def pwa_manifest_view(request, institution_slug=None):
    from django.http import JsonResponse
    inst_name = "GO FEST - Festival Portal"
    start_url = request.build_absolute_uri('/')

    if institution_slug:
        inst = Institution.objects.filter(slug=institution_slug).first()
        if inst:
            inst_name = f"{inst.name} Fest Portal"
            start_url = request.build_absolute_uri(f'/portal/{inst.slug}/contestant/dashboard/')

    data = {
        "name": inst_name,
        "short_name": "GO FEST",
        "description": f"Official Mobile App Portal for {inst_name}",
        "start_url": start_url,
        "scope": "/",
        "display": "standalone",
        "orientation": "portrait",
        "background_color": "#0f172a",
        "theme_color": "#06b6d4",
        "icons": [
            {
                "src": "/static/img/gofest_icon.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": "/static/img/GOFEST APP ICON.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ]
    }
    return JsonResponse(data, content_type="application/manifest+json")


def pwa_serviceworker_view(request):
    from django.http import HttpResponse
    js_content = """
    const CACHE_NAME = 'gofest-pwa-v1';
    const urlsToCache = [
        '/',
        '/static/css/main.css',
        '/static/img/gofest_icon.png',
        '/static/img/gf_emblem.png'
    ];

    self.addEventListener('install', (event) => {
        event.waitUntil(
            caches.open(CACHE_NAME).then((cache) => cache.addAll(urlsToCache))
        );
        self.skipWaiting();
    });

    self.addEventListener('activate', (event) => {
        event.waitUntil(self.clients.claim());
    });

    self.addEventListener('fetch', (event) => {
        event.respondWith(
            fetch(event.request).catch(() => caches.match(event.request))
        );
    });
    """
    return HttpResponse(js_content, content_type="application/javascript")


def custom_403_view(request, exception=None):
    from django.shortcuts import render
    return render(request, '403.html', status=403)


def custom_csrf_failure_view(request, reason=""):
    from django.shortcuts import render
    return render(request, '403_csrf.html', {'reason': reason}, status=403)












